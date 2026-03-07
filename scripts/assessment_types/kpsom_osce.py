"""KPSOM OSCE assessment types — I-PASS Handoff and Clinical Documentation.

Two concrete assessment types that share a common base for KPSOM-specific
data parsing (student responses with Q-number header rows, faculty scores
with merged header rows, and rubric parsing via LLM).
"""

from __future__ import annotations

import logging
import re

import pandas as pd
from openpyxl import load_workbook

from convert_rubric import convert_docx_to_text
from .base import AssessmentType, GradingResult

logger = logging.getLogger("osce_grader.kpsom")

# ---------------------------------------------------------------------------
# Section definitions with max scores
# ---------------------------------------------------------------------------

KPSOM_IPASS_SECTIONS = {
    "illness_severity": 2,
    "patient_summary": 14,
    "action_list": 5,
    "situation_awareness": 3,
    "organization": 3,
}

KPSOM_DOCUMENTATION_SECTIONS = {
    "hpi": 5,
    "social_history": 5,
    "summary_statement": 5,
    "assessment": 5,
    "plan": 5,
    "written_communication": 4,
}

# ---------------------------------------------------------------------------
# Grading prompts
# ---------------------------------------------------------------------------

CHECKLIST_GRADING_PROMPT = """\
You are a medical education expert grading a student's I-PASS handoff note \
for a Progress OSCE. You will receive the rubric criteria for one section \
and the student's response for that section.

For CHECKLIST sections (individual items worth specific points):
- Award credit for each item strictly based on the rubric criteria
- Full credit (1.0 pt per item): Criterion clearly met
- Partial credit (0.5 pt): Only where the rubric explicitly allows partial credit
- No credit (0 pts): Criterion not met
- Sum all item scores for the section total

For SCALE sections (score-level descriptors like "2: ..., 1: ..., 0: ..."):
- Match the student's response to the best-fitting score level descriptor
- Award the corresponding score as the section total

Be generous with terminology: accept clinical synonyms and paraphrasing that \
convey the same clinical concept. Do not require exact wording.

Provide a brief rationale for each scored item, then place the numeric total \
score alone on the final line with no other text."""

MILESTONE_GRADING_PROMPT = """\
You are a medical education expert grading a student's clinical documentation \
for a Progress OSCE. You will receive the milestone rubric criteria for one \
section (scored 1-5) and the student's response.

Score the response on a 1-5 integer scale using the milestone criteria. Apply \
the criteria as an experienced clinical educator would: holistically, with \
appropriate credit for correct clinical reasoning even if not verbatim. When \
in doubt between two adjacent scores, choose the higher.

Provide a one-sentence rationale, then place the integer score alone on the \
final line with no other text."""

# ---------------------------------------------------------------------------
# Rubric parsing prompts
# ---------------------------------------------------------------------------

CHECKLIST_RUBRIC_PARSE_PROMPT = """\
You are a medical education assistant parsing an OSCE grading rubric.
Extract the complete scoring criteria for each section of this I-PASS handoff rubric.

For checklist sections (Patient Summary, Action List, Situation Awareness), extract \
every individual checklist item with its point value and any partial credit conditions.
For scale sections (Illness Severity, Organization), extract the score level descriptors.

Return ONLY valid JSON:
{"section_name": {"criteria": "full criteria text with ALL items", "max_score": <number>}}

In the "criteria" field, include EVERY scorable item, one per line:
- Checklist: "- [Item description]: 1 pt (partial: 0.5 pt if [condition])\\n- [Next item]: 1 pt\\n..."
- Scale: "2: [descriptor]\\n1: [descriptor]\\n0: [descriptor]"

Sections to extract: illness_severity, patient_summary, action_list, \
situation_awareness, organization"""

MILESTONE_RUBRIC_PARSE_PROMPT = """\
You are a medical education assistant parsing an OSCE grading rubric.
Extract the milestone criteria for each section of this clinical documentation \
rubric. For each section, extract criteria at each milestone level (1-5).
Return ONLY valid JSON:
{"section_name": {"criteria": "full criteria text including all milestone \
levels", "max_score": <number>}}
Sections to extract: HPI, Social History, Summary Statement, Assessment, \
Plan, Written Communication"""


# ---------------------------------------------------------------------------
# Milestone derivation (I-PASS Handoff)
# ---------------------------------------------------------------------------

def derive_milestone(
    total: float,
    has_situation_awareness: bool,
    org_score: float,
) -> str:
    """Derive the milestone label from total score and sub-scores.

    Logic mirrors the Excel formula in the KPSOM faculty scores file.
    """
    if total > 25 and has_situation_awareness and org_score == 2:
        return "Aspirational"
    elif total > 23 and has_situation_awareness and org_score == 2:
        return "Advanced Developing to Aspirational"
    elif total > 17 and has_situation_awareness:
        return "Advanced Developing"
    elif total > 15 and has_situation_awareness:
        return "Mid-Developing to Advanced Developing"
    elif total > 13:
        return "Mid-Developing"
    elif total > 11:
        return "Early Developing to Mid-Developing"
    elif total > 8:
        return "Early Developing"
    else:
        return "Entry to Early Developing"


# ---------------------------------------------------------------------------
# Shared data-loading helpers
# ---------------------------------------------------------------------------

def _load_responses(path: str) -> pd.DataFrame:
    """Load a KPSOM student responses Excel file.

    Row 0 contains Q-number metadata (e.g. Q3, Q4, ...).
    Row 1 contains the real section name headers.
    Rows 2+ are student data.

    The Q-number row is stored as ``df.attrs['q_numbers']``.
    """
    # Read the raw file with no header to inspect both rows
    raw = pd.read_excel(path, header=None)

    if raw.shape[0] < 2:
        raise ValueError(
            f"Responses file '{path}' must have at least a Q-number row "
            "and a header row."
        )

    # Row 0 = Q-number metadata
    q_numbers = raw.iloc[0].tolist()

    # Row 1 = real column headers
    headers = raw.iloc[1].tolist()
    headers = [str(h).strip() if pd.notna(h) else f"col_{i}" for i, h in enumerate(headers)]

    # Rows 2+ = data
    df = raw.iloc[2:].reset_index(drop=True)
    df.columns = headers

    # Store Q-number mapping as metadata
    df.attrs["q_numbers"] = dict(zip(headers, q_numbers))

    return df


def _load_faculty_scores(path: str) -> pd.DataFrame:
    """Load a KPSOM faculty scores Excel file.

    Row 0–1: merged/group headers (skipped).
    Row 2: real item-level column headers.
    Rows 3+: student score data.

    Trailing rows where Student ID is not a valid integer are filtered out.
    Uses ``data_only=True`` to resolve Excel formula values.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Read all rows as lists
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 3:
        raise ValueError(
            f"Faculty scores file '{path}' must have at least 3 header rows."
        )

    # Row index 2 = real headers
    headers = [
        str(h).strip() if h is not None else f"col_{i}"
        for i, h in enumerate(all_rows[2])
    ]

    # Data rows start at index 3
    data_rows = all_rows[3:]
    df = pd.DataFrame(data_rows, columns=headers)

    # Filter to rows with valid integer Student IDs
    def _is_valid_student(val) -> bool:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return False
        try:
            int(float(val)) if isinstance(val, float) else int(val)
            return True
        except (ValueError, TypeError):
            return False

    # Find the Student column (might be "Student" or "Student ID")
    student_col = None
    for col in df.columns:
        if col.lower().strip() in ("student", "student id"):
            student_col = col
            break

    if student_col is None:
        raise ValueError(
            f"Faculty scores file '{path}' has no 'Student' or 'Student ID' "
            "column in the header row."
        )

    mask = df[student_col].apply(_is_valid_student)
    filtered = df[mask].copy()

    # Normalize Student ID to int
    filtered[student_col] = filtered[student_col].apply(
        lambda v: int(float(v)) if isinstance(v, float) else int(v)
    )

    # Warn about columns that are entirely None/NaN
    for col in filtered.columns:
        if col == student_col:
            continue
        if filtered[col].isna().all() or (filtered[col] == 0).all():
            non_none = filtered[col].notna().sum()
            if non_none == 0:
                logger.warning(
                    "Faculty scores column '%s' has no awarded values across "
                    "all %d students. This may indicate a data quality issue.",
                    col,
                    len(filtered),
                )

    return filtered


def parse_rubric_with_llm(docx_path: str, caller, task_type: str) -> dict:
    """Parse a KPSOM rubric .docx file using an LLM.

    Args:
        docx_path: Path to the rubric .docx file.
        caller: An LLMCaller instance.
        task_type: Either ``'checklist'`` or ``'milestone'``.

    Returns:
        A dict: ``{section_name: {'criteria': str, 'max_score': float}}``.
    """
    import json
    from grader import call_llm

    raw_text = convert_docx_to_text(docx_path)

    if task_type == "checklist":
        system_prompt = CHECKLIST_RUBRIC_PARSE_PROMPT
    else:
        system_prompt = MILESTONE_RUBRIC_PARSE_PROMPT

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": raw_text},
    ]

    response = call_llm(caller, messages, temperature=0.0, top_p=1.0)

    # Strip markdown code fences if present
    cleaned = response.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.split("\n", 1)[-1]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()

    parsed = json.loads(cleaned)
    return parsed


# ---------------------------------------------------------------------------
# Section name mapping helpers
# ---------------------------------------------------------------------------

# Maps common column header variations to canonical section keys
_IPASS_COLUMN_MAP = {
    "illness severity": "illness_severity",
    "patient summary": "patient_summary",
    "action list": "action_list",
    "situation awareness": "situation_awareness",
    "situational awareness": "situation_awareness",
    "situation awareness and contingency planning": "situation_awareness",
    "organization": "organization",
}

_DOCUMENTATION_COLUMN_MAP = {
    "hpi": "hpi",
    "history of present illness": "hpi",
    "social history": "social_history",
    "social hx": "social_history",
    "summary statement": "summary_statement",
    "summary": "summary_statement",
    "assessment": "assessment",
    "plan": "plan",
    "written communication": "written_communication",
    "written comm": "written_communication",
}


def _map_columns(df: pd.DataFrame, column_map: dict) -> pd.DataFrame:
    """Rename DataFrame columns using a case-insensitive mapping.

    Only renames columns that match (case-insensitive) a key in column_map.
    """
    rename = {}
    for col in df.columns:
        lower = col.lower().strip()
        if lower in column_map:
            rename[col] = column_map[lower]
    return df.rename(columns=rename)


# ---------------------------------------------------------------------------
# Score extraction helpers
# ---------------------------------------------------------------------------

def _extract_float_score(text: str) -> float | None:
    """Extract a float score from the last line of LLM output."""
    lines = text.strip().splitlines()
    for line in reversed(lines):
        line = line.strip()
        match = re.match(r"^(\d+(?:\.\d+)?)$", line)
        if match:
            return float(match.group(1))
    # Fallback: look for "Score: N" pattern
    match = re.search(r"(?:score|total)\s*[:=\-]\s*(\d+(?:\.\d+)?)", text, re.IGNORECASE)
    if match:
        return float(match.group(1))
    return None


def _extract_int_score(text: str) -> float | None:
    """Extract an integer score (1-5) from the last line of LLM output."""
    score = _extract_float_score(text)
    if score is not None:
        return float(int(score))
    return None


# ---------------------------------------------------------------------------
# KPSOM Base Type
# ---------------------------------------------------------------------------

class KPSOMBaseType(AssessmentType):
    """Shared base for KPSOM assessment types."""

    _sections: dict[str, int]       # section_name -> max_score
    _column_map: dict[str, str]     # header text -> section key
    _grading_prompt: str
    _rubric_task_type: str          # 'checklist' or 'milestone'

    def get_sections(self) -> list[str]:
        return list(self._sections.keys())

    def get_required_files(self) -> list[dict]:
        return [
            {
                "key": "rubric",
                "label": "Rubric (.docx)",
                "types": ["docx"],
                "required": True,
            },
            {
                "key": "responses",
                "label": "Student Responses (.xlsx)",
                "types": ["xlsx"],
                "required": True,
            },
            {
                "key": "scores",
                "label": "Faculty Scores (.xlsx)",
                "types": ["xlsx"],
                "required": False,
            },
        ]

    def load_inputs(self, **file_paths) -> tuple[pd.DataFrame, dict]:
        """Load student responses and optionally faculty scores.

        The rubric is NOT loaded here — it requires an LLM call and is
        handled separately via ``parse_rubric_with_llm()``.
        """
        responses_path = file_paths.get("responses")
        if not responses_path:
            raise ValueError("Student responses file is required.")

        df = _load_responses(responses_path)
        df = _map_columns(df, self._column_map)

        rubric_data: dict = {"rubric_path": file_paths.get("rubric")}

        # Pass through rubric_id for direct DB lookup (skips LLM parsing)
        if "rubric_id" in file_paths:
            rubric_data["rubric_id"] = file_paths["rubric_id"]

        scores_path = file_paths.get("scores")
        if scores_path:
            faculty_df = _load_faculty_scores(scores_path)
            # Map display-name columns to canonical section keys so that
            # process_assessment can look up scores by section key.
            faculty_df = _map_columns(faculty_df, self._column_map)
            rubric_data["faculty_scores"] = faculty_df

        return df, rubric_data

    def build_grading_prompt(self, section: str, rubric_data: dict) -> str:
        return self._grading_prompt

    def build_user_message(
        self,
        section: str,
        student_response: str,
        rubric_data: dict,
    ) -> str:
        # Include rubric criteria if available from LLM parsing
        parsed_rubric = rubric_data.get("parsed_rubric", {})
        section_rubric = parsed_rubric.get(section, {})
        criteria = section_rubric.get("criteria", "No specific criteria available.")
        max_score = section_rubric.get(
            "max_score", self._sections.get(section, "N/A")
        )

        return (
            f"Section: {section}\n"
            f"Maximum score for this section: {max_score}\n"
            f"Rubric criteria:\n{criteria}\n\n"
            f"Student response:\n{student_response}"
        )

    def build_output_df(self, results: list[GradingResult]) -> pd.DataFrame:
        rows = []
        sections = self.get_sections()

        for r in results:
            row: dict = {"student_id": r.student_id}

            # AI scores
            for sec in sections:
                row[f"{sec}_ai_score"] = r.section_scores.get(sec)
                row[f"{sec}_ai_explanation"] = r.section_explanations.get(sec, "")

            row["ai_total"] = r.total_score
            row["ai_milestone"] = r.milestone

            # Faculty scores if available
            if r.faculty_scores is not None:
                for sec in sections:
                    fac_score = r.faculty_scores.get(sec)
                    row[f"{sec}_faculty_score"] = fac_score
                    ai_score = r.section_scores.get(sec)
                    if ai_score is not None and fac_score is not None:
                        row[f"{sec}_delta"] = ai_score - fac_score
                    else:
                        row[f"{sec}_delta"] = None

                fac_total = r.faculty_scores.get("total")
                row["faculty_total"] = fac_total
                row["faculty_milestone"] = r.faculty_scores.get("milestone")
                if r.total_score is not None and fac_total is not None:
                    row["total_delta"] = r.total_score - fac_total
                else:
                    row["total_delta"] = None
                row["faculty_comments"] = r.faculty_scores.get("comments", "")

            rows.append(row)

        return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# I-PASS Handoff Type
# ---------------------------------------------------------------------------

class KPSOMHandoffType(KPSOMBaseType):
    """KPSOM OSCE — I-PASS Handoff (checklist scoring)."""

    name = "KPSOM OSCE (I-PASS Handoff)"
    type_id = "kpsom_ipass"

    _sections = KPSOM_IPASS_SECTIONS
    _column_map = _IPASS_COLUMN_MAP
    _grading_prompt = CHECKLIST_GRADING_PROMPT
    _rubric_task_type = "checklist"

    def parse_llm_response(
        self, response: str, section: str
    ) -> tuple[str, float | None]:
        score = _extract_float_score(response)
        return response, score

    def _derive_milestone_for_result(self, result: GradingResult) -> str | None:
        """Derive milestone label from a grading result."""
        if result.total_score is None:
            return None

        sa_score = result.section_scores.get("situation_awareness")
        org_score = result.section_scores.get("organization")

        if sa_score is None or org_score is None:
            return None

        has_sa = sa_score > 0
        return derive_milestone(result.total_score, has_sa, org_score)


# ---------------------------------------------------------------------------
# Clinical Documentation Type
# ---------------------------------------------------------------------------

class KPSOMDocumentationType(KPSOMBaseType):
    """KPSOM OSCE — Clinical Documentation (milestone scoring)."""

    name = "KPSOM OSCE (Clinical Documentation)"
    type_id = "kpsom_documentation"

    _sections = KPSOM_DOCUMENTATION_SECTIONS
    _column_map = _DOCUMENTATION_COLUMN_MAP
    _grading_prompt = MILESTONE_GRADING_PROMPT
    _rubric_task_type = "milestone"

    def parse_llm_response(
        self, response: str, section: str
    ) -> tuple[str, float | None]:
        score = _extract_int_score(response)
        return response, score

"""KPSOM Clinical Documentation assessment type (Case 3 pattern).

Handles clinical documentation notes with milestone-based scoring (1-5 per
section) and domain groupings (PCIG, PCDP, PCDO).

File format:
- Responses: Row 0 = section labels (single header row), Row 1+ = student data.
  Col 0 = student row number, Cols 1-5 = section text.
- Scores: Row 0 = item-level headers (flat, single header row), Row 1+ = data.
  No student ID column label — derive from row position.
  Key score columns by index: 29=HPI, 34=Social Hx, 49=Summary Statement,
  58=Assessment, 72=Plan, 73=Org/Lang, 74-81=domain totals and milestones.
"""

from __future__ import annotations

import json
import logging

import pandas as pd
from openpyxl import load_workbook

from grader import extract_score
from convert_rubric import convert_docx_to_text
from .base import AssessmentType, GradingResult
from .kpsom_osce import KPSOMBaseType, parse_rubric_with_llm

logger = logging.getLogger("osce_grader.kpsom_documentation")

# ---------------------------------------------------------------------------
# Section definitions
# ---------------------------------------------------------------------------

DOCUMENTATION_SECTIONS = {
    "hpi": 5,
    "social_hx": 5,
    "summary_statement": 5,
    "assessment": 5,
    "plan": 5,
}

# Org/Lang (max 4) is NOT AI-gradable — passthrough from faculty scores only.

# Column index -> section mapping in the responses file
_RESPONSE_COL_MAP = {
    1: "hpi",
    2: "social_hx",
    3: "summary_statement",
    4: "assessment",
    5: "plan",
}

# Score column indices in the faculty scores file (0-based)
_SCORE_COL_INDICES = {
    29: "hpi",
    34: "social_hx",
    49: "summary_statement",
    58: "assessment",
    72: "plan",
    73: "org_lang",
}

_SUMMARY_COL_INDICES = {
    74: "pcig_total",
    75: "pcig_milestone",
    76: "pcdp_total",
    77: "pcdp_milestone",
    78: "pcdo_score",
    79: "pcdo_milestone",
    80: "total_score",
    81: "total_milestone",
}

# ---------------------------------------------------------------------------
# Milestone thresholds
# ---------------------------------------------------------------------------

DOCUMENTATION_MILESTONES = [
    (0, 4, "Entry"),
    (4.5, 8, "Early Developing"),
    (8.5, 12, "Early Developing to Mid-Developing"),
    (12.5, 16, "Mid-Developing"),
    (16.5, 19, "Mid-Developing to Advanced Developing"),
    (19.5, 22, "Advanced Developing"),
    (22.5, 24, "Advanced Developing to Aspirational"),
    (24.5, 29, "Aspirational"),
]


def derive_documentation_milestone(total: float) -> str:
    """Derive milestone label from total score using documentation thresholds."""
    for low, high, label in DOCUMENTATION_MILESTONES:
        if low <= total <= high:
            return label
    # Fallback for edge cases
    if total < 0:
        return "Entry"
    return "Aspirational"


# ---------------------------------------------------------------------------
# Grading prompt
# ---------------------------------------------------------------------------

DOCUMENTATION_GRADING_PROMPT = """\
You are a medical education expert grading a student's clinical documentation \
note for a Progress OSCE. The note was written after a simulated patient \
encounter.

You will receive:
1. The rubric criteria for one section, with milestone descriptors at each \
score level (1-5)
2. The student's written response for that section

Score the response as an integer from 1 to 5 using the milestone criteria. \
Apply the criteria as an experienced clinical educator would: assess quality \
of clinical reasoning, not just keyword presence. Accept appropriate synonyms \
and paraphrasing. When the response sits between two milestone levels, choose \
the higher one.

Provide one sentence of rationale, then place the integer score alone on the \
final line with no other text."""

# ---------------------------------------------------------------------------
# Rubric parsing prompt
# ---------------------------------------------------------------------------

DOCUMENTATION_RUBRIC_PARSE_PROMPT = """\
You are a medical education assistant parsing an OSCE grading rubric.
Extract the milestone criteria for each section of this clinical documentation \
rubric. For each section, extract criteria at each milestone level (1-5).
Return ONLY valid JSON:
{"section_name": {"criteria": "full criteria text including all milestone \
levels", "max_score": <number>}}
Sections to extract: HPI, Social Hx, Summary Statement, Assessment, Plan"""

# ---------------------------------------------------------------------------
# Data loading helpers
# ---------------------------------------------------------------------------


def _load_documentation_responses(path: str) -> pd.DataFrame:
    """Load Case 3 documentation responses file.

    Row 0 = section labels (single header row). Row 1+ = student data.
    Col 0 = student row number (integer). Cols 1-5 = section text.
    """
    raw = pd.read_excel(path, header=None)

    if raw.shape[0] < 2:
        raise ValueError(
            f"Responses file '{path}' must have at least a header row "
            "and one data row."
        )

    # Row 0 is the header, rows 1+ are data
    data = raw.iloc[1:].reset_index(drop=True)

    # Build DataFrame with canonical column names
    records = []
    for _, row in data.iterrows():
        student_id = row.iloc[0]
        if pd.isna(student_id):
            continue
        try:
            student_id = int(float(student_id))
        except (ValueError, TypeError):
            continue

        record = {"student_id": student_id}
        for col_idx, section_name in _RESPONSE_COL_MAP.items():
            if col_idx < len(row):
                val = row.iloc[col_idx]
                if pd.notna(val) and str(val).strip():
                    record[section_name] = str(val)
                else:
                    record[section_name] = None
            else:
                record[section_name] = None
        records.append(record)

    df = pd.DataFrame(records)
    return df


def _load_documentation_scores(path: str) -> dict:
    """Load Case 3 documentation faculty scores file.

    Row 0 = item-level headers (flat). Row 1+ = student data.
    No student ID column — derive from row position (1-indexed).
    Returns dict keyed by student_id (int).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) < 2:
        raise ValueError(
            f"Scores file '{path}' must have at least a header row and one data row."
        )

    # Row 0 = headers, Row 1+ = data
    data_rows = all_rows[1:]
    result = {}

    for row_idx, row in enumerate(data_rows):
        student_id = row_idx + 1  # 1-indexed

        scores = {}
        # Section scores
        for col_idx, section_name in _SCORE_COL_INDICES.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    try:
                        scores[section_name] = float(val)
                    except (ValueError, TypeError):
                        scores[section_name] = None
                else:
                    scores[section_name] = None
            else:
                scores[section_name] = None

        # Summary columns
        for col_idx, key in _SUMMARY_COL_INDICES.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    if "milestone" in key:
                        scores[key] = str(val)
                    else:
                        try:
                            scores[key] = float(val)
                        except (ValueError, TypeError):
                            scores[key] = None
                else:
                    scores[key] = None
            else:
                scores[key] = None

        result[student_id] = scores

    return result


# ---------------------------------------------------------------------------
# KPSOMDocumentationType
# ---------------------------------------------------------------------------


class KPSOMDocumentationType(AssessmentType):
    """KPSOM OSCE — Clinical Documentation (milestone scoring, Case 3)."""

    name = "KPSOM OSCE (Clinical Documentation)"
    type_id = "kpsom_documentation"

    _sections = DOCUMENTATION_SECTIONS
    _rubric_task_type = "documentation"
    _rubric_parse_prompt = DOCUMENTATION_RUBRIC_PARSE_PROMPT

    def get_sections(self) -> list[str]:
        return list(self._sections.keys())

    def get_required_files(self) -> list[dict]:
        return [
            {
                "key": "rubric",
                "label": "Rubric (.docx)",
                "types": ["docx"],
                "required": True,
            },
            {
                "key": "responses",
                "label": "Student Responses (.xlsx)",
                "types": ["xlsx"],
                "required": True,
            },
            {
                "key": "scores",
                "label": "Faculty Scores (.xlsx)",
                "types": ["xlsx"],
                "required": False,
            },
        ]

    def load_inputs(self, **file_paths) -> tuple[pd.DataFrame, dict]:
        """Load student responses and optionally faculty scores."""
        responses_path = file_paths.get("responses")
        if not responses_path:
            raise ValueError("Student responses file is required.")

        df = _load_documentation_responses(responses_path)
        rubric_data: dict = {"rubric_path": file_paths.get("rubric")}
        if "rubric_id" in file_paths:
            rubric_data["rubric_id"] = file_paths["rubric_id"]

        scores_path = file_paths.get("scores")
        if scores_path:
            rubric_data["faculty_scores_dict"] = _load_documentation_scores(
                scores_path
            )

        return df, rubric_data

    def build_grading_prompt(self, section: str, rubric_data: dict) -> str:
        return DOCUMENTATION_GRADING_PROMPT

    def build_user_message(
        self,
        section: str,
        student_response: str,
        rubric_data: dict,
    ) -> str:
        parsed_rubric = rubric_data.get("parsed_rubric", {})
        section_rubric = parsed_rubric.get(section, {})
        criteria = section_rubric.get(
            "criteria", "No specific criteria available."
        )
        max_score = section_rubric.get(
            "max_score", self._sections.get(section, "N/A")
        )

        return (
            f"Section: {section.upper().replace('_', ' ')}\n"
            f"Rubric criteria:\n{criteria}\n\n"
            f"Student response:\n{student_response}\n\n"
            f"Score (1\u2013{max_score}):"
        )

    def parse_llm_response(
        self, response: str, section: str
    ) -> tuple[str, float | None]:
        score = extract_score(response)
        if score is not None:
            max_score = self._sections.get(section, 5)
            score = max(1, min(score, max_score))
            return response, float(score)
        return response, None

    def build_output_df(self, results: list[GradingResult]) -> pd.DataFrame:
        rows = []
        sections = self.get_sections()

        for r in results:
            row: dict = {"student_id": r.student_id}

            # AI scores per section
            for sec in sections:
                row[f"{sec}_ai_score"] = r.section_scores.get(sec)

            # Domain totals
            hpi = r.section_scores.get("hpi")
            social_hx = r.section_scores.get("social_hx")
            summary = r.section_scores.get("summary_statement")
            assessment = r.section_scores.get("assessment")
            plan = r.section_scores.get("plan")

            pcig = None
            if hpi is not None and social_hx is not None:
                pcig = hpi + social_hx
            row["ai_pcig_total"] = pcig

            pcdp = None
            if all(v is not None for v in [summary, assessment, plan]):
                pcdp = summary + assessment + plan
            row["ai_pcdp_total"] = pcdp

            ai_total = None
            if pcig is not None and pcdp is not None:
                ai_total = pcig + pcdp
            row["ai_total"] = ai_total

            row["ai_total_milestone"] = (
                derive_documentation_milestone(ai_total)
                if ai_total is not None
                else None
            )

            # Explanations
            for sec in sections:
                row[f"{sec}_ai_explanation"] = r.section_explanations.get(
                    sec, ""
                )

            # Faculty scores if available
            if r.faculty_scores is not None:
                for sec in sections:
                    row[f"{sec}_faculty_score"] = r.faculty_scores.get(sec)
                    ai_score = r.section_scores.get(sec)
                    fac_score = r.faculty_scores.get(sec)
                    if ai_score is not None and fac_score is not None:
                        row[f"{sec}_delta"] = ai_score - fac_score
                    else:
                        row[f"{sec}_delta"] = None

                row["org_lang_faculty_score"] = r.faculty_scores.get("org_lang")
                row["faculty_pcig_total"] = r.faculty_scores.get("pcig_total")
                row["faculty_pcdp_total"] = r.faculty_scores.get("pcdp_total")
                row["faculty_pcdo_score"] = r.faculty_scores.get("pcdo_score")
                row["faculty_total"] = r.faculty_scores.get("total_score")
                row["faculty_total_milestone"] = r.faculty_scores.get(
                    "total_milestone"
                )

                fac_total = r.faculty_scores.get("total_score")
                if ai_total is not None and fac_total is not None:
                    row["total_delta"] = ai_total - fac_total
                else:
                    row["total_delta"] = None

                row["faculty_comments"] = r.faculty_scores.get("comments", "")

            rows.append(row)

        return pd.DataFrame(rows)

    def _derive_milestone_for_result(
        self, result: GradingResult
    ) -> str | None:
        """Derive milestone from AI total score."""
        if result.total_score is None:
            return None
        return derive_documentation_milestone(result.total_score)

"""KPSOM Ethics OSCE assessment type (Case 8 pattern).

Handles open-ended ethics questions with sub-scored components:
- Q1: dual sub-score (Problematic + Reasonable, max 4)
- Q2A/B/C: pro/con analysis (max 2 each, total Q2 max 6)
- Q3: four capacity questions (max 2 each, total max 8)
- Total: 18 points

File format:
- Responses: Row 0 = full question text headers. Col 0 = student ID.
  Q2 has separate Pro/Con columns with redundant combined columns to skip.
- Scores: Row 0 = section groups, Row 1 = sub-item labels, Row 2+ = data.
"""

from __future__ import annotations

import json
import logging
import re

import pandas as pd
from openpyxl import load_workbook

from .base import AssessmentType, GradingResult
from .kpsom_osce import KPSOMBaseType

logger = logging.getLogger("osce_grader.kpsom_ethics")

# ---------------------------------------------------------------------------
# Section definitions
# ---------------------------------------------------------------------------

ETHICS_SECTIONS = {
    "q1": 4,
    "q2a": 2,
    "q2b": 2,
    "q2c": 2,
    "q3": 8,
}

# ---------------------------------------------------------------------------
# Milestone thresholds (exact from rubric scoring table)
# ---------------------------------------------------------------------------

ETHICS_MILESTONES = [
    (0, 3, "Entry"),
    (3.5, 5.5, "Entry to Early Developing"),
    (6.0, 7.5, "Early Developing"),
    (8, 9.5, "Early Developing to Mid-Developing"),
    (10, 11.5, "Mid-Developing"),
    (12, 13, "Mid-Developing to Advanced Developing"),
    (13.5, 15, "Advanced Developing"),
    (15.5, 16, "Advanced Developing to Aspirational"),
    (16.5, 18, "Aspirational"),
]


def derive_ethics_milestone(total: float) -> str:
    """Derive milestone label from total score using ethics thresholds."""
    for low, high, label in ETHICS_MILESTONES:
        if low <= total <= high:
            return label
    if total < 0:
        return "Entry"
    return "Aspirational"


# ---------------------------------------------------------------------------
# Grading prompts
# ---------------------------------------------------------------------------

Q1_GRADING_PROMPT = """\
You are a medical education expert grading a student's response to an ethics \
question about informed consent in a Progress OSCE.

The question has two parts:
1. Identify what is problematic about asking a student to consent a patient \
(up to 2 points — 1 point per clearly identified problem, max 2)
2. Describe how the student would respond (up to 2 points — 1 point per \
reasonable element of what they would do, max 2)

Evaluate the student response and return ONLY valid JSON:
{"q1_problematic_score": <0, 1, or 2>, "q1_reasonable_score": <0, 1, or 2>, \
"rationale": "<one sentence>"}"""

Q2_GRADING_PROMPT = """\
You are a medical education expert grading a student's pros and cons analysis \
for one option in an informed consent scenario for a Progress OSCE.

Award 1 point if the student identified at least one reasonable pro, 0 if not.
Award 1 point if the student identified at least one reasonable con, 0 if not.
Be generous: accept any clinically or ethically reasonable pro/con, not just \
those listed in the rubric.

Return ONLY valid JSON:
{"score": <0, 0.5, 1, 1.5, or 2>, "rationale": "<one sentence>"}"""

Q3_GRADING_PROMPT = """\
You are a medical education expert grading a student's four questions about \
decision-making capacity for a Progress OSCE.

The 5 key elements of decision-making capacity are:
1. Appreciates current relevant medical situation
2. Understands the intervention proposed, with risks and benefits
3. Understands alternatives to proposed intervention, with risks and benefits
4. Is able to express or communicate a choice
5. Is able to give reasons for the choice

Scoring rules:
- 2 points: question clearly and well addresses one of the 5 key elements \
above, asked in plain patient-friendly language
- 1 point: question addresses a key element but is asked poorly, unclearly, \
or uses jargon the patient might not understand
- 0 points: question addresses orientation, diagnosis, mental status, or other \
issues NOT on the key element list — even if clinically relevant
- No double-counting: if 2 or more questions address the same element, score \
only the best of them; the others score 0
- Maximum 8 points total across 4 questions

Return ONLY valid JSON:
{"scores": [<score for Q1>, <score for Q2>, <score for Q3>, <score for Q4>], \
"elements_addressed": ["element name or 'orientation/other'", ...], \
"rationale": "<one sentence>"}"""

# ---------------------------------------------------------------------------
# Rubric parsing prompt
# ---------------------------------------------------------------------------

ETHICS_RUBRIC_PARSE_PROMPT = """\
You are a medical education assistant. Parse this ethics OSCE rubric and extract \
the scoring criteria for each question component.

Return ONLY valid JSON with these keys:
{
  "q1_problematic": "criteria for what makes asking a student to consent a \
patient problematic (up to 2 points)",
  "q1_reasonable": "criteria for what a reasonable student response would be \
(up to 2 points)",
  "q2a": "criteria for pros and cons of calling the daughter to sign (up to \
2 points total)",
  "q2b": "criteria for pros and cons of performing MSE and asking patient to \
sign (up to 2 points total)",
  "q2c": "criteria for pros and cons of asking a senior resident to handle \
consent (up to 2 points total)",
  "q3": "criteria for the 5 key elements of decision-making capacity and how \
questions are scored (2 pts each, max 8)"
}"""

# ---------------------------------------------------------------------------
# Data loading helpers
# ---------------------------------------------------------------------------


def _load_ethics_responses(path: str) -> pd.DataFrame:
    """Load Case 8 ethics responses file.

    Row 0 = column headers (full question text). Row 1+ = student data.
    Col 0 = student_id. Col 1 = Q1. Cols 2,3 = Q2A Pro/Con.
    Cols 5,6 = Q2B Pro/Con. Cols 8,9 = Q2C Pro/Con.
    Cols 11-14 = Q3 questions 1-4.
    Cols 4, 7, 10 = redundant combined fields (skipped).
    """
    raw = pd.read_excel(path, header=None)

    if raw.shape[0] < 2:
        raise ValueError(
            f"Responses file '{path}' must have at least a header row "
            "and one data row."
        )

    # Row 0 = headers (store for reference), rows 1+ = data
    data = raw.iloc[1:].reset_index(drop=True)
    records = []

    for _, row in data.iterrows():
        student_id = row.iloc[0]
        if pd.isna(student_id):
            continue
        try:
            student_id = int(float(student_id))
        except (ValueError, TypeError):
            continue

        # Check Q1 for "No PET submitted" or empty
        q1_raw = row.iloc[1] if len(row) > 1 else None
        if q1_raw is None or (
            isinstance(q1_raw, str)
            and (
                not q1_raw.strip()
                or "no pet submitted" in q1_raw.lower()
            )
        ) or (not isinstance(q1_raw, str) and pd.isna(q1_raw)):
            # Non-submitter — include with None values
            records.append({
                "student_id": student_id,
                "q1": None,
                "q2a": None,
                "q2b": None,
                "q2c": None,
                "q3": None,
                "_is_submitter": False,
            })
            continue

        record = {"student_id": student_id, "_is_submitter": True}

        # Q1
        record["q1"] = str(q1_raw) if pd.notna(q1_raw) else None

        # Q2A: concatenate Pro (col 2) and Con (col 3)
        q2a_pro = row.iloc[2] if len(row) > 2 else None
        q2a_con = row.iloc[3] if len(row) > 3 else None
        record["q2a"] = _concat_pro_con(q2a_pro, q2a_con)

        # Q2B: Pro (col 5), Con (col 6)
        q2b_pro = row.iloc[5] if len(row) > 5 else None
        q2b_con = row.iloc[6] if len(row) > 6 else None
        record["q2b"] = _concat_pro_con(q2b_pro, q2b_con)

        # Q2C: Pro (col 8), Con (col 9)
        q2c_pro = row.iloc[8] if len(row) > 8 else None
        q2c_con = row.iloc[9] if len(row) > 9 else None
        record["q2c"] = _concat_pro_con(q2c_pro, q2c_con)

        # Q3: concatenate cols 11-14 as numbered list
        q3_parts = []
        for i, col_idx in enumerate([11, 12, 13, 14], start=1):
            if col_idx < len(row):
                val = row.iloc[col_idx]
                if pd.notna(val) and str(val).strip():
                    q3_parts.append(f"{i}. {str(val).strip()}")
        record["q3"] = "\n".join(q3_parts) if q3_parts else None

        records.append(record)

    df = pd.DataFrame(records)
    return df


def _concat_pro_con(pro, con) -> str | None:
    """Concatenate pro and con values into a single string."""
    parts = []
    if pro is not None and pd.notna(pro) and str(pro).strip():
        parts.append(f"Pro: {str(pro).strip()}")
    if con is not None and pd.notna(con) and str(con).strip():
        parts.append(f"Con: {str(con).strip()}")
    return "\n".join(parts) if parts else None


def _load_ethics_scores(path: str) -> dict:
    """Load Case 8 ethics faculty scores file.

    Row 0 = section group labels. Row 1 = sub-item labels. Row 2+ = data.
    Returns dict keyed by student_id (int).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) < 3:
        raise ValueError(
            f"Scores file '{path}' must have at least two header rows "
            "and one data row."
        )

    # Row 0 and 1 are headers; Row 2+ is data
    data_rows = all_rows[2:]
    result = {}

    for row in data_rows:
        student_id = row[0]
        if student_id is None or (
            isinstance(student_id, float) and pd.isna(student_id)
        ):
            continue
        try:
            student_id = int(float(student_id)) if isinstance(
                student_id, float
            ) else int(student_id)
        except (ValueError, TypeError):
            continue

        def _safe_float(val):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        scores = {
            "q1_problematic": _safe_float(row[2]) if len(row) > 2 else None,
            "q1_reasonable": _safe_float(row[3]) if len(row) > 3 else None,
            "q1_total": _safe_float(row[4]) if len(row) > 4 else None,
            "q2a_score": _safe_float(row[6]) if len(row) > 6 else None,
            "q2b_score": _safe_float(row[8]) if len(row) > 8 else None,
            "q2c_score": _safe_float(row[10]) if len(row) > 10 else None,
            "q2_total": _safe_float(row[11]) if len(row) > 11 else None,
            "q3_total": _safe_float(row[16]) if len(row) > 16 else None,
            "task_total": _safe_float(row[17]) if len(row) > 17 else None,
            "milestone": (
                str(row[18]) if len(row) > 18 and row[18] is not None else None
            ),
            "comments": (
                str(row[19]) if len(row) > 19 and row[19] is not None else None
            ),
        }
        result[student_id] = scores

    return result


# ---------------------------------------------------------------------------
# JSON parsing helpers
# ---------------------------------------------------------------------------


def _strip_json_fences(text: str) -> str:
    """Strip markdown code fences from LLM JSON output."""
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.split("\n", 1)[-1]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()
    return cleaned


def _parse_q1_response(response: str) -> tuple[str, float | None]:
    """Parse Q1 JSON response with two sub-scores."""
    try:
        parsed = json.loads(_strip_json_fences(response))
        prob = float(parsed.get("q1_problematic_score", 0))
        reas = float(parsed.get("q1_reasonable_score", 0))
        prob = max(0, min(2, prob))
        reas = max(0, min(2, reas))
        total = prob + reas
        rationale = parsed.get("rationale", "")

        # Encode sub-scores as JSON metadata in explanation
        explanation = json.dumps({
            "q1_problematic_score": prob,
            "q1_reasonable_score": reas,
            "rationale": rationale,
        })
        return explanation, total
    except (json.JSONDecodeError, TypeError, KeyError):
        logger.warning("Failed to parse Q1 JSON response, attempting fallback.")
        # Fallback: try to extract a total score from text
        from grader import extract_score

        score = extract_score(response)
        if score is not None:
            return response, min(float(score), 4.0)
        return response, None


def _parse_q2_response(response: str) -> tuple[str, float | None]:
    """Parse Q2 JSON response with single score."""
    try:
        parsed = json.loads(_strip_json_fences(response))
        score = float(parsed.get("score", 0))
        score = max(0, min(2, score))
        rationale = parsed.get("rationale", "")

        explanation = json.dumps({
            "score": score,
            "rationale": rationale,
        })
        return explanation, score
    except (json.JSONDecodeError, TypeError, KeyError):
        logger.warning("Failed to parse Q2 JSON response, attempting fallback.")
        from grader import extract_score

        score = extract_score(response)
        if score is not None:
            return response, min(float(score), 2.0)
        return response, None


def _parse_q3_response(response: str) -> tuple[str, float | None]:
    """Parse Q3 JSON response with per-question scores."""
    try:
        parsed = json.loads(_strip_json_fences(response))
        scores = parsed.get("scores", [])
        if not isinstance(scores, list) or len(scores) != 4:
            raise ValueError("Expected 4 scores")
        scores = [max(0, min(2, float(s))) for s in scores]
        total = sum(scores)
        total = min(8, total)
        elements = parsed.get("elements_addressed", [])
        rationale = parsed.get("rationale", "")

        explanation = json.dumps({
            "scores": scores,
            "elements_addressed": elements,
            "rationale": rationale,
        })
        return explanation, total
    except (json.JSONDecodeError, TypeError, KeyError, ValueError):
        logger.warning("Failed to parse Q3 JSON response, attempting fallback.")
        from grader import extract_score

        score = extract_score(response)
        if score is not None:
            return response, min(float(score), 8.0)
        return response, None


# ---------------------------------------------------------------------------
# KPSOMEthicsType
# ---------------------------------------------------------------------------


class KPSOMEthicsType(AssessmentType):
    """KPSOM OSCE — Ethics Open-Ended Questions (Case 8)."""

    name = "KPSOM OSCE (Ethics)"
    type_id = "kpsom_ethics"

    _sections = ETHICS_SECTIONS
    _rubric_task_type = "ethics"
    _rubric_parse_prompt = ETHICS_RUBRIC_PARSE_PROMPT

    def get_sections(self) -> list[str]:
        return list(self._sections.keys())

    def get_required_files(self) -> list[dict]:
        return [
            {
                "key": "rubric",
                "label": "Rubric (.docx)",
                "types": ["docx"],
                "required": True,
            },
            {
                "key": "responses",
                "label": "Student Responses (.xlsx)",
                "types": ["xlsx"],
                "required": True,
            },
            {
                "key": "scores",
                "label": "Faculty Scores (.xlsx)",
                "types": ["xlsx"],
                "required": False,
            },
        ]

    def load_inputs(self, **file_paths) -> tuple[pd.DataFrame, dict]:
        """Load student responses and optionally faculty scores."""
        responses_path = file_paths.get("responses")
        if not responses_path:
            raise ValueError("Student responses file is required.")

        df = _load_ethics_responses(responses_path)
        rubric_data: dict = {"rubric_path": file_paths.get("rubric")}
        if "rubric_id" in file_paths:
            rubric_data["rubric_id"] = file_paths["rubric_id"]

        scores_path = file_paths.get("scores")
        if scores_path:
            rubric_data["faculty_scores_dict"] = _load_ethics_scores(scores_path)

        return df, rubric_data

    def build_grading_prompt(self, section: str, rubric_data: dict) -> str:
        if section == "q1":
            parsed = rubric_data.get("parsed_rubric", {})
            q1_prob = parsed.get("q1_problematic", "Award 1 point per clearly identified problem, max 2")
            q1_reas = parsed.get("q1_reasonable", "Award 1 point per reasonable response element, max 2")
            return (
                Q1_GRADING_PROMPT.rstrip()
                + f"\n\nRubric \u2014 Problematic aspects (award 1 pt each, max 2):\n{q1_prob}"
                + f"\n\nRubric \u2014 Reasonable response elements (award 1 pt each, max 2):\n{q1_reas}"
            )
        elif section.startswith("q2"):
            parsed = rubric_data.get("parsed_rubric", {})
            section_criteria = parsed.get(section, "1 point for a reasonable pro, 1 point for a reasonable con")
            return (
                Q2_GRADING_PROMPT.rstrip()
                + f"\n\nThis option's rubric:\n{section_criteria}"
            )
        elif section == "q3":
            return Q3_GRADING_PROMPT
        return Q1_GRADING_PROMPT

    def build_user_message(
        self,
        section: str,
        student_response: str,
        rubric_data: dict,
    ) -> str:
        return f"Student response:\n{student_response}"

    def parse_llm_response(
        self, response: str, section: str
    ) -> tuple[str, float | None]:
        if section == "q1":
            return _parse_q1_response(response)
        elif section.startswith("q2"):
            return _parse_q2_response(response)
        elif section == "q3":
            return _parse_q3_response(response)
        return response, None

    def build_output_df(self, results: list[GradingResult]) -> pd.DataFrame:
        rows = []
        sections = self.get_sections()

        for r in results:
            row: dict = {"student_id": r.student_id}

            # Q1 sub-scores
            q1_explanation = r.section_explanations.get("q1", "")
            q1_prob = None
            q1_reas = None
            try:
                q1_data = json.loads(q1_explanation)
                q1_prob = q1_data.get("q1_problematic_score")
                q1_reas = q1_data.get("q1_reasonable_score")
            except (json.JSONDecodeError, TypeError):
                pass

            q1_total = r.section_scores.get("q1")
            row["q1_ai_total"] = q1_total
            row["q1_ai_problematic"] = q1_prob
            row["q1_ai_reasonable"] = q1_reas

            # Q2 scores
            q2a = r.section_scores.get("q2a")
            q2b = r.section_scores.get("q2b")
            q2c = r.section_scores.get("q2c")
            row["q2a_ai_score"] = q2a
            row["q2b_ai_score"] = q2b
            row["q2c_ai_score"] = q2c

            q2_total = None
            if all(v is not None for v in [q2a, q2b, q2c]):
                q2_total = q2a + q2b + q2c
            row["q2_ai_total"] = q2_total

            # Q3 score
            row["q3_ai_total"] = r.section_scores.get("q3")

            # Total
            ai_total = r.total_score
            row["ai_total"] = ai_total
            row["ai_milestone"] = (
                derive_ethics_milestone(ai_total)
                if ai_total is not None
                else None
            )

            # Explanations
            for sec in sections:
                row[f"{sec}_ai_explanation"] = r.section_explanations.get(
                    sec, ""
                )

            # Faculty scores if available
            if r.faculty_scores is not None:
                fac = r.faculty_scores
                row["q1_faculty_total"] = fac.get("q1_total")
                row["q1_faculty_problematic"] = fac.get("q1_problematic")
                row["q1_faculty_reasonable"] = fac.get("q1_reasonable")
                row["q2a_faculty_score"] = fac.get("q2a_score")
                row["q2b_faculty_score"] = fac.get("q2b_score")
                row["q2c_faculty_score"] = fac.get("q2c_score")
                row["q2_faculty_total"] = fac.get("q2_total")
                row["q3_faculty_total"] = fac.get("q3_total")
                row["faculty_total"] = fac.get("task_total")
                row["faculty_milestone"] = fac.get("milestone")

                # Deltas
                if q1_total is not None and fac.get("q1_total") is not None:
                    row["q1_delta"] = q1_total - fac["q1_total"]
                else:
                    row["q1_delta"] = None

                if q2_total is not None and fac.get("q2_total") is not None:
                    row["q2_delta"] = q2_total - fac["q2_total"]
                else:
                    row["q2_delta"] = None

                q3_ai = r.section_scores.get("q3")
                if q3_ai is not None and fac.get("q3_total") is not None:
                    row["q3_delta"] = q3_ai - fac["q3_total"]
                else:
                    row["q3_delta"] = None

                fac_total = fac.get("task_total")
                if ai_total is not None and fac_total is not None:
                    row["total_delta"] = ai_total - fac_total
                else:
                    row["total_delta"] = None

                row["faculty_comments"] = fac.get("comments", "")

            rows.append(row)

        return pd.DataFrame(rows)

    def _derive_milestone_for_result(
        self, result: GradingResult
    ) -> str | None:
        """Derive milestone from AI total score."""
        if result.total_score is None:
            return None
        return derive_ethics_milestone(result.total_score)

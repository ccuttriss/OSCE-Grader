"""Gold Standard Faculty Bias Analysis.

Aggregates faculty scores from multiple administrations of the same
assessment, runs statistical and LLM-powered bias analysis, and
produces a gold standard benchmark artifact for faculty calibration.
"""

from __future__ import annotations

import io
import json
import logging
import statistics
from dataclasses import dataclass, field
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class SessionData:
    """Faculty scores from one administration of an assessment."""

    label: str
    assessment_type_id: str
    sections: list[str]
    scores: dict[int, dict]  # student_id → {section: score, …}
    student_count: int


@dataclass
class SectionStats:
    """Aggregated statistics for a single section across sessions."""

    section: str
    mean: float
    median: float
    std: float
    min_score: float
    max_score: float
    per_session_means: dict[str, float] = field(default_factory=dict)
    per_session_stds: dict[str, float] = field(default_factory=dict)


@dataclass
class CrossSessionStats:
    """Aggregated statistics across all uploaded sessions."""

    assessment_type_id: str
    sections: list[str]
    section_stats: dict[str, SectionStats] = field(default_factory=dict)
    total_students: int = 0
    session_count: int = 0
    session_labels: list[str] = field(default_factory=list)


@dataclass
class BiasAnalysisResult:
    """Structured output from LLM bias analysis."""

    systematic_biases: list[dict] = field(default_factory=list)
    distribution_anomalies: list[dict] = field(default_factory=list)
    section_patterns: list[dict] = field(default_factory=list)
    drift_over_years: list[dict] = field(default_factory=list)
    outlier_patterns: list[dict] = field(default_factory=list)
    summary: str = ""
    recommendations: list[str] = field(default_factory=list)


@dataclass
class ConsensusResult:
    """Output from Cultural Consensus Theory analysis."""

    eigenvalues: list[float]
    eigenvalue_ratio: float                    # λ₁/λ₂
    single_culture_holds: bool                 # True if ratio >= 3.0
    fit_label: str                             # "Strong" / "Adequate" / "Weak" / etc.
    session_competence: dict[str, float]       # session_label → normalized weight
    first_factor_loadings: dict[str, float]    # session_label → signed loading
    consensus_means: dict[str, float]          # section → competence-weighted mean
    simple_means: dict[str, float]             # section → unweighted mean
    divergence: dict[str, float]               # section → |consensus - simple|
    has_negative_loadings: bool = False


@dataclass
class GoldStandardBenchmark:
    """Complete gold standard artifact ready for export."""

    assessment_type_id: str
    created_date: str
    sections: dict[str, dict] = field(default_factory=dict)
    bias_findings: BiasAnalysisResult = field(default_factory=BiasAnalysisResult)
    stats: CrossSessionStats | None = None
    consensus: ConsensusResult | None = None
    version: int = 1


# ---------------------------------------------------------------------------
# Section key mappings per assessment type
# ---------------------------------------------------------------------------

# Which keys in the faculty score dicts represent gradable section scores
# (as opposed to totals, milestones, comments, etc.)
_SECTION_KEYS: dict[str, list[str]] = {
    "kpsom_ipass": [
        "illness_severity",
        "patient_summary",
        "action_list",
        "situation_awareness",
        "organization",
    ],
    "kpsom_documentation": [
        "hpi",
        "social_hx",
        "summary_statement",
        "assessment",
        "plan",
        "org_lang",
    ],
    "kpsom_ethics": [
        "q1_total",
        "q2a_score",
        "q2b_score",
        "q2c_score",
        "q3_total",
    ],
    "uk_osce": [
        "hpi",
        "pex",
        "sum",
        "ddx",
        "support",
        "plan",
    ],
}


# ---------------------------------------------------------------------------
# Session loading
# ---------------------------------------------------------------------------


def load_faculty_session(
    path: str,
    assessment_type_id: str,
    label: str,
) -> SessionData:
    """Load a single faculty score file and return a SessionData.

    Delegates to the appropriate loader based on *assessment_type_id*.
    """
    sections = _SECTION_KEYS.get(assessment_type_id)
    if sections is None:
        raise ValueError(f"Unknown assessment type: {assessment_type_id}")

    if assessment_type_id == "kpsom_ipass":
        scores = _load_ipass_session(path)
    elif assessment_type_id == "kpsom_documentation":
        scores = _load_documentation_session(path)
    elif assessment_type_id == "kpsom_ethics":
        scores = _load_ethics_session(path)
    elif assessment_type_id == "uk_osce":
        scores = _load_uk_osce_session(path)
    else:
        raise ValueError(f"Unknown assessment type: {assessment_type_id}")

    return SessionData(
        label=label,
        assessment_type_id=assessment_type_id,
        sections=list(sections),
        scores=scores,
        student_count=len(scores),
    )


def _load_ipass_session(path: str) -> dict[int, dict]:
    """Load I-PASS handoff faculty scores, returning dict keyed by student ID."""
    from assessment_types.kpsom_osce import _load_faculty_scores

    df = _load_faculty_scores(path)

    # Find the Student column
    student_col = None
    for col in df.columns:
        if col.lower().strip() in ("student", "student id"):
            student_col = col
            break
    if student_col is None:
        raise ValueError("No Student/Student ID column found")

    sections = _SECTION_KEYS["kpsom_ipass"]
    result: dict[int, dict] = {}

    # Map DataFrame columns to section keys (case-insensitive)
    col_map = {}
    for col in df.columns:
        lower = col.lower().strip()
        from assessment_types.kpsom_osce import _IPASS_COLUMN_MAP

        if lower in _IPASS_COLUMN_MAP:
            col_map[col] = _IPASS_COLUMN_MAP[lower]

    for _, row in df.iterrows():
        sid = int(row[student_col])
        scores: dict[str, float | None] = {}
        for orig_col, section_key in col_map.items():
            if section_key in sections:
                val = row[orig_col]
                if pd.notna(val):
                    try:
                        scores[section_key] = float(val)
                    except (ValueError, TypeError):
                        scores[section_key] = None
                else:
                    scores[section_key] = None
        result[sid] = scores

    return result


def _load_documentation_session(path: str) -> dict[int, dict]:
    """Load documentation faculty scores, filtering to section-level keys."""
    from assessment_types.kpsom_documentation import _load_documentation_scores

    raw = _load_documentation_scores(path)
    sections = _SECTION_KEYS["kpsom_documentation"]
    return {
        sid: {k: v for k, v in sdict.items() if k in sections}
        for sid, sdict in raw.items()
    }


def _load_ethics_session(path: str) -> dict[int, dict]:
    """Load ethics faculty scores, filtering to section-level keys."""
    from assessment_types.kpsom_ethics import _load_ethics_scores

    raw = _load_ethics_scores(path)
    sections = _SECTION_KEYS["kpsom_ethics"]
    return {
        sid: {k: v for k, v in sdict.items() if k in sections}
        for sid, sdict in raw.items()
    }


def _load_uk_osce_session(path: str) -> dict[int, dict]:
    """Load UK OSCE faculty scores from a simple flat Excel file.

    Expects columns matching section names (hpi, pex, sum, ddx, support, plan)
    with one row per student. First column or a 'student_id' column provides IDs.
    """
    df = pd.read_excel(path)
    sections = _SECTION_KEYS["uk_osce"]

    # Find student ID column
    sid_col = None
    for col in df.columns:
        if col.lower().strip() in ("student_id", "student", "id"):
            sid_col = col
            break

    result: dict[int, dict] = {}
    for row_idx, row in df.iterrows():
        sid = int(row[sid_col]) if sid_col else row_idx + 1
        scores: dict[str, float | None] = {}
        for sec in sections:
            # Try exact match then case-insensitive
            val = None
            if sec in df.columns:
                val = row[sec]
            else:
                for col in df.columns:
                    if col.lower().strip() == sec:
                        val = row[col]
                        break
            if val is not None and pd.notna(val):
                try:
                    scores[sec] = float(val)
                except (ValueError, TypeError):
                    scores[sec] = None
            else:
                scores[sec] = None
        result[sid] = scores

    return result


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def validate_sessions(sessions: list[SessionData]) -> str | None:
    """Validate that all sessions are compatible.

    Returns an error message string or ``None`` if valid.
    """
    if len(sessions) < 2:
        return "At least 2 sessions are required for cross-session analysis."
    if len(sessions) > 10:
        return "Maximum of 10 sessions allowed."

    type_ids = {s.assessment_type_id for s in sessions}
    if len(type_ids) > 1:
        return (
            f"All sessions must be the same assessment type. "
            f"Found: {', '.join(sorted(type_ids))}"
        )

    return None


# ---------------------------------------------------------------------------
# Statistical analysis
# ---------------------------------------------------------------------------


def compute_cross_session_stats(
    sessions: list[SessionData],
) -> CrossSessionStats:
    """Compute per-section statistics aggregated across all sessions."""
    if not sessions:
        raise ValueError("No sessions provided.")

    type_id = sessions[0].assessment_type_id
    sections = sessions[0].sections

    stats = CrossSessionStats(
        assessment_type_id=type_id,
        sections=list(sections),
        session_count=len(sessions),
        session_labels=[s.label for s in sessions],
    )

    total_students = 0
    for session in sessions:
        total_students += session.student_count
    stats.total_students = total_students

    for section in sections:
        all_values: list[float] = []
        per_session_means: dict[str, float] = {}
        per_session_stds: dict[str, float] = {}

        for session in sessions:
            session_values: list[float] = []
            for _sid, sdict in session.scores.items():
                val = sdict.get(section)
                if val is not None:
                    session_values.append(val)
                    all_values.append(val)

            if session_values:
                per_session_means[session.label] = statistics.mean(session_values)
                per_session_stds[session.label] = (
                    statistics.stdev(session_values)
                    if len(session_values) > 1
                    else 0.0
                )

        if all_values:
            sec_stats = SectionStats(
                section=section,
                mean=statistics.mean(all_values),
                median=statistics.median(all_values),
                std=(
                    statistics.stdev(all_values)
                    if len(all_values) > 1
                    else 0.0
                ),
                min_score=min(all_values),
                max_score=max(all_values),
                per_session_means=per_session_means,
                per_session_stds=per_session_stds,
            )
        else:
            sec_stats = SectionStats(
                section=section,
                mean=0.0,
                median=0.0,
                std=0.0,
                min_score=0.0,
                max_score=0.0,
            )

        stats.section_stats[section] = sec_stats

    return stats


# ---------------------------------------------------------------------------
# Cultural Consensus Analysis (CCT)
# ---------------------------------------------------------------------------

logger = logging.getLogger(__name__)


def _build_session_section_matrix(
    stats: CrossSessionStats,
) -> tuple[np.ndarray, list[str], list[str]]:
    """Build the sessions × sections matrix of per-session mean scores.

    Returns (matrix, session_labels, section_names).
    """
    labels = stats.session_labels
    sections = stats.sections

    matrix = np.zeros((len(labels), len(sections)))
    for i, label in enumerate(labels):
        for j, sec in enumerate(sections):
            ss = stats.section_stats.get(sec)
            if ss is None:
                matrix[i, j] = np.nan
            else:
                val = ss.per_session_means.get(label)
                matrix[i, j] = val if val is not None else np.nan

    # Validate — no entirely-missing row or column
    for i, label in enumerate(labels):
        if np.all(np.isnan(matrix[i, :])):
            raise ValueError(
                f"Session '{label}' has no scores for any section."
            )
    for j, sec in enumerate(sections):
        if np.all(np.isnan(matrix[:, j])):
            raise ValueError(
                f"Section '{sec}' has no scores in any session."
            )

    return matrix, labels, sections


def _compute_agreement_matrix(matrix: np.ndarray) -> np.ndarray:
    """Compute the Pearson correlation matrix across sessions (rows).

    Returns an n_sessions × n_sessions correlation matrix.
    """
    # Z-score each column (section) to normalize across different scales
    col_means = np.nanmean(matrix, axis=0)
    col_stds = np.nanstd(matrix, axis=0, ddof=0)
    # Avoid division by zero for constant columns
    col_stds[col_stds < 1e-10] = 1.0
    standardized = (matrix - col_means) / col_stds

    # Replace any remaining NaN with 0 (neutral contribution)
    standardized = np.nan_to_num(standardized, nan=0.0)

    # Check for degenerate case: if all rows are identical (perfect agreement),
    # corrcoef will produce NaN. Detect this and return all-ones matrix.
    row_stds = np.std(standardized, axis=1)
    if np.all(row_stds < 1e-10):
        n = standardized.shape[0]
        return np.ones((n, n))

    # Compute correlation of rows (sessions) across columns (sections)
    agreement = np.corrcoef(standardized)

    # Handle NaN from constant rows (session with zero variance after z-scoring)
    agreement = np.nan_to_num(agreement, nan=0.0)

    return agreement


def _eigendecompose(
    agreement: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    """Eigendecompose the agreement matrix.

    Returns (eigenvalues_descending, eigenvectors_corresponding).
    Eigenvectors are columns of the returned matrix.
    """
    eigenvalues, eigenvectors = np.linalg.eigh(agreement)
    # eigh returns ascending order — reverse to descending
    idx = np.argsort(eigenvalues)[::-1]
    eigenvalues = eigenvalues[idx]
    eigenvectors = eigenvectors[:, idx]
    return eigenvalues, eigenvectors


def compute_consensus_analysis(
    sessions: list[SessionData],
    stats: CrossSessionStats,
) -> ConsensusResult:
    """Run Cultural Consensus Theory analysis on cross-session data.

    Treats each session as an "informant" and each section as an "item".
    Uses the informal model (Pearson correlations + eigendecomposition)
    to estimate session competence and a consensus-weighted answer key.
    """
    if len(sessions) < 2:
        raise ValueError("At least 2 sessions required for consensus analysis.")

    # Step 1: Build the sessions × sections matrix
    matrix, labels, sections = _build_session_section_matrix(stats)
    n_sessions = len(labels)

    # Step 2: Compute agreement matrix (correlations across sessions)
    agreement = _compute_agreement_matrix(matrix)

    # Step 3: Eigendecompose
    eigenvalues, eigenvectors = _eigendecompose(agreement)

    # Step 4: Eigenvalue ratio and fit assessment
    lambda1 = float(eigenvalues[0])
    lambda2 = float(eigenvalues[1]) if n_sessions > 1 else 0.0

    if abs(lambda2) < 1e-10:
        eigenvalue_ratio = float("inf")
    else:
        eigenvalue_ratio = lambda1 / abs(lambda2)

    if n_sessions == 2:
        fit_label = "N/A (2 sessions)"
        single_culture_holds = eigenvalue_ratio >= 3.0
    elif eigenvalue_ratio >= 5.0:
        fit_label = "Strong"
        single_culture_holds = True
    elif eigenvalue_ratio >= 3.0:
        fit_label = "Adequate"
        single_culture_holds = True
    else:
        fit_label = "Weak"
        single_culture_holds = False

    # Step 5: First-factor loadings and competence weights
    first_eigenvector = eigenvectors[:, 0]
    has_negative = bool(np.any(first_eigenvector < 0))

    # Signed loadings (for display — user wants to see these)
    first_factor_loadings = {}
    for i, label in enumerate(labels):
        first_factor_loadings[label] = float(first_eigenvector[i])

    # Competence weights = |loading| normalized to sum to 1
    abs_loadings = np.abs(first_eigenvector)
    total = abs_loadings.sum()
    if total < 1e-10:
        # Degenerate — equal weights
        weights = np.ones(n_sessions) / n_sessions
    else:
        weights = abs_loadings / total

    session_competence = {}
    for i, label in enumerate(labels):
        session_competence[label] = float(weights[i])

    # Step 6: Consensus answer key (competence-weighted means)
    # Use the ORIGINAL (unstandardized) per-session means
    consensus_means = {}
    simple_means = {}
    divergence = {}

    for j, sec in enumerate(sections):
        sec_values = matrix[:, j]  # Original means per session
        # Weighted mean
        valid_mask = ~np.isnan(sec_values)
        if valid_mask.any():
            valid_vals = sec_values[valid_mask]
            valid_weights = weights[valid_mask]
            w_sum = valid_weights.sum()
            if w_sum > 0:
                consensus_means[sec] = float(
                    np.sum(valid_vals * valid_weights) / w_sum
                )
            else:
                consensus_means[sec] = float(np.mean(valid_vals))
            simple_means[sec] = float(np.mean(valid_vals))
        else:
            consensus_means[sec] = 0.0
            simple_means[sec] = 0.0

        divergence[sec] = abs(consensus_means[sec] - simple_means[sec])

    return ConsensusResult(
        eigenvalues=[float(v) for v in eigenvalues],
        eigenvalue_ratio=eigenvalue_ratio,
        single_culture_holds=single_culture_holds,
        fit_label=fit_label,
        session_competence=session_competence,
        first_factor_loadings=first_factor_loadings,
        consensus_means=consensus_means,
        simple_means=simple_means,
        divergence=divergence,
        has_negative_loadings=has_negative,
    )


# ---------------------------------------------------------------------------
# LLM bias analysis
# ---------------------------------------------------------------------------

_TYPE_DISPLAY_NAMES = {
    "uk_osce": "Standard OSCE (Post-Encounter Notes)",
    "kpsom_ipass": "KPSOM I-PASS Handoff",
    "kpsom_documentation": "KPSOM Clinical Documentation",
    "kpsom_ethics": "KPSOM Ethics Open-Ended Questions",
}

BIAS_ANALYSIS_SYSTEM_PROMPT = """\
You are a psychometrician and medical education assessment expert. You are \
analyzing faculty scoring data from multiple administrations of the same \
{assessment_type_name} assessment to identify potential scoring biases and \
quality concerns.

Analyze the provided scoring statistics and identify:

1. SYSTEMATIC BIASES: Sections where faculty scores are consistently higher \
or lower than expected given the rubric scale, suggesting leniency or severity \
bias. Look for means that cluster near the top or bottom of the scale.

2. DISTRIBUTION ANOMALIES: Sections with unusual score distributions (e.g., \
very low standard deviation suggesting rubber-stamping, very high std suggesting \
inconsistent application of criteria, ceiling or floor effects).

3. SECTION PATTERNS: Sections that are consistently scored differently from \
others across all sessions (e.g., Plan scores always lowest), which may \
indicate the rubric is unclear or faculty interpret criteria inconsistently.

4. DRIFT OVER YEARS: Changes in scoring patterns across sessions that may \
indicate grading standards shifting over time. Compare per-session means for \
each section.

5. OUTLIER PATTERNS: Recurring patterns of extreme scores that may indicate \
systematic issues rather than individual student performance.

Be specific and evidence-based. Reference actual numbers from the data. \
Distinguish between statistically meaningful patterns and normal variation. \
Consider the scoring scale when interpreting magnitudes.

Respond with ONLY valid JSON in this format:
{{
  "systematic_biases": [{{"section": "...", "direction": "lenient|severe", \
"magnitude": "low|medium|high", "description": "..."}}],
  "distribution_anomalies": [{{"section": "...", "description": "..."}}],
  "section_patterns": [{{"section": "...", "pattern": "...", \
"description": "..."}}],
  "drift_over_years": [{{"section": "...", "direction": "increasing|decreasing|\
variable", "description": "..."}}],
  "outlier_patterns": [{{"description": "..."}}],
  "summary": "2-3 sentence overall assessment of scoring quality and bias risks",
  "recommendations": ["actionable recommendation 1", "..."]
}}
"""


def build_bias_prompt(
    stats: CrossSessionStats,
    type_id: str,
) -> list[dict[str, str]]:
    """Build system + user messages for the LLM bias analysis."""
    type_name = _TYPE_DISPLAY_NAMES.get(type_id, type_id)

    system = BIAS_ANALYSIS_SYSTEM_PROMPT.format(assessment_type_name=type_name)

    # Build the user message with formatted statistics
    lines = [
        f"Assessment Type: {type_name}",
        f"Sessions: {stats.session_count} ({', '.join(stats.session_labels)})",
        f"Total Students: {stats.total_students}",
        "",
        "=== PER-SECTION STATISTICS ===",
    ]

    for section in stats.sections:
        ss = stats.section_stats.get(section)
        if ss is None:
            continue
        lines.append(f"\nSection: {section}")
        lines.append(
            f"  Overall: mean={ss.mean:.2f}, median={ss.median:.2f}, "
            f"std={ss.std:.2f}, range=[{ss.min_score:.1f}-{ss.max_score:.1f}]"
        )
        for label in stats.session_labels:
            s_mean = ss.per_session_means.get(label)
            s_std = ss.per_session_stds.get(label)
            if s_mean is not None:
                lines.append(
                    f"  {label}: mean={s_mean:.2f}"
                    + (f", std={s_std:.2f}" if s_std is not None else "")
                )

    # Cross-section comparison
    section_means = {
        sec: ss.mean
        for sec, ss in stats.section_stats.items()
        if ss.mean > 0
    }
    if section_means:
        lowest = min(section_means, key=section_means.get)
        highest = max(section_means, key=section_means.get)
        lines.extend([
            "",
            "=== CROSS-SECTION COMPARISON ===",
            f"Lowest-scored section: {lowest} (mean={section_means[lowest]:.2f})",
            f"Highest-scored section: {highest} (mean={section_means[highest]:.2f})",
        ])

        # Highest variance section
        section_stds = {
            sec: ss.std
            for sec, ss in stats.section_stats.items()
            if ss.std > 0
        }
        if section_stds:
            most_variable = max(section_stds, key=section_stds.get)
            lines.append(
                f"Most variable section: {most_variable} "
                f"(std={section_stds[most_variable]:.2f})"
            )

    return [
        {"role": "system", "content": system},
        {"role": "user", "content": "\n".join(lines)},
    ]


def parse_bias_response(text: str) -> BiasAnalysisResult:
    """Parse the LLM JSON response into a BiasAnalysisResult.

    Falls back to an empty result with the raw text as summary on parse failure.
    """
    cleaned = _strip_json_fences(text)
    try:
        data = json.loads(cleaned)
    except (json.JSONDecodeError, ValueError):
        return BiasAnalysisResult(
            summary=f"Failed to parse LLM response as JSON. Raw output:\n{text}"
        )

    return BiasAnalysisResult(
        systematic_biases=data.get("systematic_biases", []),
        distribution_anomalies=data.get("distribution_anomalies", []),
        section_patterns=data.get("section_patterns", []),
        drift_over_years=data.get("drift_over_years", []),
        outlier_patterns=data.get("outlier_patterns", []),
        summary=data.get("summary", ""),
        recommendations=data.get("recommendations", []),
    )


def _strip_json_fences(text: str) -> str:
    """Strip markdown code fences from LLM JSON output."""
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.split("\n", 1)[-1]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()
    return cleaned


# ---------------------------------------------------------------------------
# Benchmark artifact generation
# ---------------------------------------------------------------------------


def generate_benchmark_excel(benchmark: GoldStandardBenchmark) -> bytes:
    """Generate a multi-sheet Excel workbook for the gold standard benchmark.

    Sheet 1: Per-section benchmarks (min, max, notes, approver)
    Sheet 2: Bias findings
    Sheet 3: Cross-session statistics
    """
    wb = Workbook()

    # --- Sheet 1: Benchmarks ---
    ws1 = wb.active
    ws1.title = "Benchmarks"
    ws1.append([
        "Section",
        "Benchmark Min",
        "Benchmark Max",
        "Notes",
        "Approved By",
    ])
    for section, info in benchmark.sections.items():
        ws1.append([
            section,
            info.get("min"),
            info.get("max"),
            info.get("notes", ""),
            info.get("approved_by", ""),
        ])

    # --- Sheet 2: Bias Findings ---
    ws2 = wb.create_sheet("Bias Findings")
    ws2.append(["Category", "Section", "Detail"])

    bf = benchmark.bias_findings
    for item in bf.systematic_biases:
        ws2.append([
            "Systematic Bias",
            item.get("section", ""),
            f"[{item.get('direction', '')}] {item.get('description', '')}",
        ])
    for item in bf.distribution_anomalies:
        ws2.append([
            "Distribution Anomaly",
            item.get("section", ""),
            item.get("description", ""),
        ])
    for item in bf.section_patterns:
        ws2.append([
            "Section Pattern",
            item.get("section", ""),
            item.get("description", ""),
        ])
    for item in bf.drift_over_years:
        ws2.append([
            "Year-over-Year Drift",
            item.get("section", ""),
            item.get("description", ""),
        ])
    for item in bf.outlier_patterns:
        ws2.append([
            "Outlier Pattern",
            "",
            item.get("description", ""),
        ])

    ws2.append([])
    ws2.append(["Summary", "", bf.summary])
    for rec in bf.recommendations:
        ws2.append(["Recommendation", "", rec])

    # --- Sheet 3: Statistics ---
    ws3 = wb.create_sheet("Statistics")

    if benchmark.stats:
        st = benchmark.stats
        ws3.append([
            "Section",
            "Mean",
            "Median",
            "Std Dev",
            "Min",
            "Max",
        ] + [f"{lbl} Mean" for lbl in st.session_labels])

        for section in st.sections:
            ss = st.section_stats.get(section)
            if ss is None:
                continue
            row = [
                section,
                round(ss.mean, 2),
                round(ss.median, 2),
                round(ss.std, 2),
                round(ss.min_score, 2),
                round(ss.max_score, 2),
            ]
            for lbl in st.session_labels:
                m = ss.per_session_means.get(lbl)
                row.append(round(m, 2) if m is not None else "")
            ws3.append(row)

    # --- Sheet 4: Consensus Analysis ---
    if benchmark.consensus:
        ws4 = wb.create_sheet("Consensus Analysis")
        c = benchmark.consensus

        # Model fit
        ws4.append(["Cultural Consensus Theory Analysis"])
        ws4.append([])
        ws4.append(["Eigenvalue Ratio", round(c.eigenvalue_ratio, 2)])
        ws4.append(["Model Fit", c.fit_label])
        ws4.append([
            "Single Culture",
            "Yes" if c.single_culture_holds else "No",
        ])
        ws4.append([
            "Negative Loadings",
            "Yes" if c.has_negative_loadings else "No",
        ])

        # Eigenvalues
        ws4.append([])
        ws4.append(["Eigenvalues"])
        ws4.append(["Index", "Value"])
        for i, ev in enumerate(c.eigenvalues):
            ws4.append([i + 1, round(ev, 4)])

        # Session competence
        ws4.append([])
        ws4.append(["Session Competence Scores"])
        ws4.append(["Session", "First-Factor Loading (signed)", "Weight (normalized)"])
        for label in c.session_competence:
            ws4.append([
                label,
                round(c.first_factor_loadings[label], 4),
                round(c.session_competence[label], 4),
            ])

        # Consensus answer key
        ws4.append([])
        ws4.append(["Consensus Answer Key"])
        ws4.append(["Section", "Consensus Mean", "Simple Mean", "Difference"])
        for sec in c.consensus_means:
            ws4.append([
                sec,
                round(c.consensus_means[sec], 3),
                round(c.simple_means[sec], 3),
                round(c.divergence[sec], 3),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_benchmark_json(benchmark: GoldStandardBenchmark) -> str:
    """Generate a machine-readable JSON representation of the benchmark."""
    data = {
        "version": benchmark.version,
        "assessment_type_id": benchmark.assessment_type_id,
        "created_date": benchmark.created_date,
        "sections": benchmark.sections,
        "bias_findings": {
            "systematic_biases": benchmark.bias_findings.systematic_biases,
            "distribution_anomalies": benchmark.bias_findings.distribution_anomalies,
            "section_patterns": benchmark.bias_findings.section_patterns,
            "drift_over_years": benchmark.bias_findings.drift_over_years,
            "outlier_patterns": benchmark.bias_findings.outlier_patterns,
            "summary": benchmark.bias_findings.summary,
            "recommendations": benchmark.bias_findings.recommendations,
        },
    }

    if benchmark.stats:
        st = benchmark.stats
        data["statistics"] = {
            "session_count": st.session_count,
            "session_labels": st.session_labels,
            "total_students": st.total_students,
            "sections": {},
        }
        for sec, ss in st.section_stats.items():
            data["statistics"]["sections"][sec] = {
                "mean": round(ss.mean, 3),
                "median": round(ss.median, 3),
                "std": round(ss.std, 3),
                "min": round(ss.min_score, 3),
                "max": round(ss.max_score, 3),
                "per_session_means": {
                    k: round(v, 3) for k, v in ss.per_session_means.items()
                },
            }

    if benchmark.consensus:
        c = benchmark.consensus
        data["consensus"] = {
            "eigenvalues": [round(v, 4) for v in c.eigenvalues],
            "eigenvalue_ratio": round(c.eigenvalue_ratio, 3),
            "single_culture_holds": c.single_culture_holds,
            "fit_label": c.fit_label,
            "has_negative_loadings": c.has_negative_loadings,
            "first_factor_loadings": {
                k: round(v, 4) for k, v in c.first_factor_loadings.items()
            },
            "session_competence": {
                k: round(v, 4) for k, v in c.session_competence.items()
            },
            "consensus_means": {
                k: round(v, 3) for k, v in c.consensus_means.items()
            },
            "simple_means": {
                k: round(v, 3) for k, v in c.simple_means.items()
            },
            "divergence": {
                k: round(v, 3) for k, v in c.divergence.items()
            },
        }

    return json.dumps(data, indent=2)

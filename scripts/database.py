"""SQLite persistence layer for OSCE Grader.

Replaces flat-file storage (synthetic_data/, examples/, .env) with a single
SQLite database for structured data management at school scale.
"""

from __future__ import annotations

import json
import logging
import os
import sqlite3
import tempfile
from contextlib import contextmanager
from typing import Optional

logger = logging.getLogger("osce_grader.database")

def _default_db_path() -> str:
    import server_env
    return server_env.db_path()

DB_PATH = _default_db_path()

CURRENT_SCHEMA_VERSION = 6

# ---------------------------------------------------------------------------
# Connection management
# ---------------------------------------------------------------------------

@contextmanager
def get_connection():
    """Yield a SQLite connection with WAL mode and foreign keys enabled."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Schema creation
# ---------------------------------------------------------------------------

_SCHEMA_SQL = """\
CREATE TABLE IF NOT EXISTS schema_version (
    version    INTEGER NOT NULL,
    applied_at TEXT    NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS api_keys (
    provider   TEXT PRIMARY KEY,
    env_var    TEXT NOT NULL,
    value      TEXT NOT NULL,
    updated_at TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS rubrics (
    rubric_id            INTEGER PRIMARY KEY AUTOINCREMENT,
    assessment_type_id   TEXT NOT NULL,
    case_title           TEXT NOT NULL DEFAULT '',
    case_description     TEXT NOT NULL DEFAULT '',
    learner_instructions TEXT NOT NULL DEFAULT '',
    model_answer         TEXT NOT NULL DEFAULT '',
    score_table_json     TEXT NOT NULL DEFAULT '[]',
    source               TEXT NOT NULL DEFAULT 'synthetic',
    created_at           TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS rubric_sections (
    rubric_section_id    INTEGER PRIMARY KEY AUTOINCREMENT,
    rubric_id            INTEGER NOT NULL REFERENCES rubrics(rubric_id) ON DELETE CASCADE,
    section_key          TEXT NOT NULL,
    display_name         TEXT NOT NULL,
    max_score            REAL NOT NULL,
    criteria             TEXT NOT NULL DEFAULT '',
    score_levels_json    TEXT NOT NULL DEFAULT '{}',
    checklist_items_json TEXT NOT NULL DEFAULT '[]',
    sort_order           INTEGER NOT NULL DEFAULT 0,
    UNIQUE(rubric_id, section_key)
);

CREATE TABLE IF NOT EXISTS synthetic_sessions (
    session_id             INTEGER PRIMARY KEY AUTOINCREMENT,
    assessment_type_id     TEXT NOT NULL,
    rubric_id              INTEGER NOT NULL REFERENCES rubrics(rubric_id) ON DELETE CASCADE,
    label                  TEXT NOT NULL,
    faculty_json           TEXT NOT NULL,
    students_json          TEXT NOT NULL,
    student_notes_json     TEXT NOT NULL,
    faculty_scores_json    TEXT NOT NULL,
    sections_json          TEXT NOT NULL,
    model_used             TEXT,
    generation_params_json TEXT DEFAULT '{}',
    is_active              INTEGER NOT NULL DEFAULT 1,
    created_at             TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS example_files (
    example_id         INTEGER PRIMARY KEY AUTOINCREMENT,
    assessment_type_id TEXT NOT NULL,
    slot               TEXT NOT NULL,
    original_filename  TEXT NOT NULL,
    file_data          BLOB NOT NULL,
    uploaded_at        TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(assessment_type_id, slot)
);

CREATE TABLE IF NOT EXISTS grading_runs (
    run_id             INTEGER PRIMARY KEY AUTOINCREMENT,
    assessment_type_id TEXT NOT NULL,
    rubric_id          INTEGER REFERENCES rubrics(rubric_id),
    model_used         TEXT NOT NULL,
    temperature        REAL NOT NULL,
    source_type        TEXT NOT NULL DEFAULT 'uploaded',
    session_id         INTEGER REFERENCES synthetic_sessions(session_id),
    started_at         TEXT NOT NULL DEFAULT (datetime('now')),
    completed_at       TEXT,
    status             TEXT NOT NULL DEFAULT 'in_progress',
    results_json       TEXT,
    log_text           TEXT
);

CREATE TABLE IF NOT EXISTS audit_events (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    ts              TEXT    NOT NULL DEFAULT (datetime('now')),
    stream          TEXT    NOT NULL CHECK (stream IN ('user','system')),
    severity        TEXT    NOT NULL CHECK (severity IN ('info','warn','error')),
    action          TEXT    NOT NULL,
    actor_email     TEXT,
    actor_role      TEXT,
    session_id      TEXT,
    request_id      TEXT,
    target_kind     TEXT,
    target_id       TEXT,
    target_hash     TEXT,
    outcome         TEXT    NOT NULL CHECK (outcome IN ('success','failure','denied')),
    error_code      TEXT,
    detail_json     TEXT
);
CREATE INDEX IF NOT EXISTS idx_audit_ts            ON audit_events(ts);
CREATE INDEX IF NOT EXISTS idx_audit_actor         ON audit_events(actor_email, ts);
CREATE INDEX IF NOT EXISTS idx_audit_stream_action ON audit_events(stream, action, ts);
CREATE INDEX IF NOT EXISTS idx_audit_target        ON audit_events(target_kind, target_id);

CREATE TABLE IF NOT EXISTS materials (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    kind            TEXT    NOT NULL CHECK (kind IN (
                        'rubric','answer_key','student_notes','exemplar')),
    display_name    TEXT    NOT NULL,
    filename        TEXT    NOT NULL,
    content_sha256  TEXT    NOT NULL,
    size_bytes      INTEGER NOT NULL,
    mime_type       TEXT,
    assessment_type TEXT,
    uploaded_by     TEXT    NOT NULL,
    uploaded_at     TEXT    NOT NULL DEFAULT (datetime('now')),
    archived_at     TEXT,
    notes           TEXT,
    UNIQUE (content_sha256, kind)
);
CREATE INDEX IF NOT EXISTS idx_materials_kind     ON materials(kind);
CREATE INDEX IF NOT EXISTS idx_materials_assess   ON materials(assessment_type);
CREATE INDEX IF NOT EXISTS idx_materials_archived ON materials(archived_at);

CREATE TABLE IF NOT EXISTS material_tags (
    material_id INTEGER NOT NULL REFERENCES materials(id) ON DELETE CASCADE,
    tag         TEXT    NOT NULL,
    PRIMARY KEY (material_id, tag)
);

CREATE TABLE IF NOT EXISTS model_config (
    model_name         TEXT PRIMARY KEY,
    provider           TEXT NOT NULL,
    enabled            INTEGER NOT NULL DEFAULT 0,
    accuracy           REAL,
    cost_1k            REAL,
    bias               REAL,
    last_benchmarked   TEXT,
    benchmark_run_id   INTEGER,
    updated_at         TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS benchmark_runs (
    run_id             INTEGER PRIMARY KEY AUTOINCREMENT,
    started_at         TEXT NOT NULL DEFAULT (datetime('now')),
    completed_at       TEXT,
    status             TEXT NOT NULL DEFAULT 'in_progress',
    sample_source      TEXT NOT NULL,
    sample_description TEXT,
    models_json        TEXT NOT NULL DEFAULT '[]',
    results_json       TEXT,
    error_text         TEXT
);
CREATE INDEX IF NOT EXISTS idx_benchmark_runs_started ON benchmark_runs(started_at);

CREATE TABLE IF NOT EXISTS users (
    email                TEXT PRIMARY KEY,
    password_hash        TEXT NOT NULL DEFAULT '',
    must_change_password INTEGER NOT NULL DEFAULT 1,
    role                 TEXT NOT NULL DEFAULT 'end_user'
                              CHECK (role IN ('admin', 'end_user')),
    created_at           TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at           TEXT NOT NULL DEFAULT (datetime('now'))
);
"""

# Bootstrap admin seeded on first init. Password is blank and must be set
# by the operator on initial sign-in.
BOOTSTRAP_ADMIN_EMAIL = "kpsomit@kp.org"


def _apply_migration_v4(conn: sqlite3.Connection) -> None:
    cols = [r[1] for r in conn.execute("PRAGMA table_info(grading_runs)").fetchall()]
    adds = [
        ("run_uuid",                "TEXT"),
        ("user_email",              "TEXT"),
        ("auth_session_id",         "TEXT"),
        ("provider",                "TEXT"),
        ("top_p",                   "REAL"),
        ("workers",                 "INTEGER"),
        ("max_tokens",              "INTEGER"),
        ("sections_json",           "TEXT"),
        ("rubric_material_id",      "INTEGER"),
        ("answer_key_material_id",  "INTEGER"),
        ("student_notes_sha256",    "TEXT"),
        ("summary_json",            "TEXT"),
        ("results_sha256",          "TEXT"),
    ]
    for name, typ in adds:
        if name not in cols:
            conn.execute(f"ALTER TABLE grading_runs ADD COLUMN {name} {typ}")
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_grading_runs_run_uuid "
        "ON grading_runs(run_uuid) WHERE run_uuid IS NOT NULL"
    )


def init_db() -> None:
    """Create tables if they don't exist. Idempotent."""
    with get_connection() as conn:
        conn.executescript(_SCHEMA_SQL)
        current_row = conn.execute(
            "SELECT COALESCE(MAX(version), 0) FROM schema_version"
        ).fetchone()
        current = current_row[0]
        if current < 4:
            _apply_migration_v4(conn)

        # Seed the bootstrap admin on first init (or any init where the
        # users table is empty). Password is blank; the operator is forced
        # to set one on first sign-in.
        has_users = conn.execute(
            "SELECT 1 FROM users LIMIT 1"
        ).fetchone() is not None
        if not has_users:
            conn.execute(
                "INSERT OR IGNORE INTO users "
                "(email, password_hash, must_change_password, role) "
                "VALUES (?, '', 1, 'admin')",
                (BOOTSTRAP_ADMIN_EMAIL,),
            )
            logger.info("Seeded bootstrap admin user %s", BOOTSTRAP_ADMIN_EMAIL)

        if current < CURRENT_SCHEMA_VERSION:
            conn.execute(
                "INSERT INTO schema_version (version) VALUES (?)",
                (CURRENT_SCHEMA_VERSION,),
            )
            logger.info(
                "Database migrated from v%d to v%d", current, CURRENT_SCHEMA_VERSION,
            )
            try:
                from audit import log_event
                log_event(
                    "db.migration",
                    stream="system",
                    severity="info",
                    detail={"from": current, "to": CURRENT_SCHEMA_VERSION},
                )
            except Exception:
                # Audit emission is best-effort during bootstrap; swallow if
                # the audit module isn't importable (e.g., during schema init
                # itself before tables exist on a fresh DB). log_event already
                # has its own never-raise guard, but this catches import errors.
                pass


# ---------------------------------------------------------------------------
# API Keys
# ---------------------------------------------------------------------------

def save_api_key(provider: str, env_var: str, value: str) -> None:
    """Store or update an API key. Value is encrypted at rest when possible."""
    import key_vault
    stored = key_vault.encrypt(value)
    with get_connection() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO api_keys (provider, env_var, value, updated_at) "
            "VALUES (?, ?, ?, datetime('now'))",
            (provider, env_var, stored),
        )


def load_api_keys() -> dict[str, tuple[str, str]]:
    """Return ``{provider: (env_var, value)}`` for all stored keys.

    Values that were written in an older plaintext format are returned as-is.
    Encrypted values are decrypted transparently.
    """
    import key_vault
    with get_connection() as conn:
        rows = conn.execute("SELECT provider, env_var, value FROM api_keys").fetchall()
    result: dict[str, tuple[str, str]] = {}
    for r in rows:
        try:
            result[r["provider"]] = (r["env_var"], key_vault.decrypt(r["value"]))
        except ValueError as exc:
            logger.warning("Skipping undecryptable key for %s: %s", r["provider"], exc)
    return result


def get_api_key(provider: str) -> Optional[str]:
    """Return the decrypted key value for a provider, or None."""
    import key_vault
    with get_connection() as conn:
        row = conn.execute(
            "SELECT value FROM api_keys WHERE provider = ?", (provider,)
        ).fetchone()
    if row is None:
        return None
    try:
        return key_vault.decrypt(row["value"])
    except ValueError as exc:
        logger.warning("Could not decrypt stored key for %s: %s", provider, exc)
        return None


def delete_api_key(provider: str) -> None:
    """Remove the stored key for a provider. No-op if absent."""
    with get_connection() as conn:
        conn.execute("DELETE FROM api_keys WHERE provider = ?", (provider,))


# ---------------------------------------------------------------------------
# Model configuration (enabled/disabled state + cached benchmark metrics)
# ---------------------------------------------------------------------------

def upsert_model_config(
    model_name: str,
    provider: str,
    *,
    enabled: Optional[bool] = None,
    accuracy: Optional[float] = None,
    cost_1k: Optional[float] = None,
    bias: Optional[float] = None,
    last_benchmarked: Optional[str] = None,
    benchmark_run_id: Optional[int] = None,
) -> None:
    """Insert or update a row in model_config.

    Only fields passed explicitly (non-None) are updated on an existing row.
    On insert, unspecified fields use table defaults.
    """
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT model_name FROM model_config WHERE model_name = ?",
            (model_name,),
        ).fetchone()

        if existing is None:
            conn.execute(
                "INSERT INTO model_config "
                "(model_name, provider, enabled, accuracy, cost_1k, bias, "
                " last_benchmarked, benchmark_run_id) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    model_name,
                    provider,
                    1 if enabled else 0,
                    accuracy,
                    cost_1k,
                    bias,
                    last_benchmarked,
                    benchmark_run_id,
                ),
            )
            return

        updates: list[str] = ["provider = ?", "updated_at = datetime('now')"]
        values: list = [provider]
        if enabled is not None:
            updates.append("enabled = ?")
            values.append(1 if enabled else 0)
        if accuracy is not None:
            updates.append("accuracy = ?")
            values.append(accuracy)
        if cost_1k is not None:
            updates.append("cost_1k = ?")
            values.append(cost_1k)
        if bias is not None:
            updates.append("bias = ?")
            values.append(bias)
        if last_benchmarked is not None:
            updates.append("last_benchmarked = ?")
            values.append(last_benchmarked)
        if benchmark_run_id is not None:
            updates.append("benchmark_run_id = ?")
            values.append(benchmark_run_id)

        values.append(model_name)
        conn.execute(
            f"UPDATE model_config SET {', '.join(updates)} WHERE model_name = ?",
            values,
        )


def set_model_enabled(model_name: str, enabled: bool) -> None:
    """Flip the enabled flag for an existing model_config row."""
    with get_connection() as conn:
        conn.execute(
            "UPDATE model_config SET enabled = ?, updated_at = datetime('now') "
            "WHERE model_name = ?",
            (1 if enabled else 0, model_name),
        )


def list_model_configs() -> list[dict]:
    """Return all model_config rows as a list of dicts."""
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM model_config ORDER BY provider, model_name"
        ).fetchall()
    return [dict(r) for r in rows]


def get_model_config(model_name: str) -> Optional[dict]:
    """Return one model_config row as a dict, or None."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM model_config WHERE model_name = ?", (model_name,)
        ).fetchone()
    return dict(row) if row else None


# ---------------------------------------------------------------------------
# Benchmark runs
# ---------------------------------------------------------------------------

def create_benchmark_run(
    sample_source: str,
    sample_description: str,
    models: list[str],
) -> int:
    """Create a benchmark_runs row in 'in_progress' state. Return run_id."""
    with get_connection() as conn:
        cur = conn.execute(
            "INSERT INTO benchmark_runs "
            "(sample_source, sample_description, models_json, status) "
            "VALUES (?, ?, ?, 'in_progress')",
            (sample_source, sample_description, json.dumps(models)),
        )
        return cur.lastrowid


def complete_benchmark_run(
    run_id: int,
    results: list[dict],
    *,
    status: str = "complete",
    error_text: Optional[str] = None,
) -> None:
    """Mark a benchmark run as finished and store its per-model results."""
    with get_connection() as conn:
        conn.execute(
            "UPDATE benchmark_runs SET completed_at = datetime('now'), "
            "status = ?, results_json = ?, error_text = ? WHERE run_id = ?",
            (status, json.dumps(results), error_text, run_id),
        )


def list_benchmark_runs(limit: int = 20) -> list[dict]:
    """Return recent benchmark runs, newest first."""
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT run_id, started_at, completed_at, status, sample_source, "
            "sample_description, models_json "
            "FROM benchmark_runs ORDER BY run_id DESC LIMIT ?",
            (limit,),
        ).fetchall()
    return [dict(r) for r in rows]


def get_benchmark_run(run_id: int) -> Optional[dict]:
    """Return a full benchmark run row as a dict, or None."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM benchmark_runs WHERE run_id = ?", (run_id,)
        ).fetchone()
    return dict(row) if row else None


# ---------------------------------------------------------------------------
# Users (authentication)
# ---------------------------------------------------------------------------

def get_user(email: str) -> Optional[dict]:
    """Return the user row for *email* as a dict, or None."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM users WHERE email = ?", (email.lower(),)
        ).fetchone()
    return dict(row) if row else None


def list_users() -> list[dict]:
    """Return every user row, newest first."""
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT email, role, must_change_password, created_at, updated_at "
            "FROM users ORDER BY created_at DESC"
        ).fetchall()
    return [dict(r) for r in rows]


def create_user(
    email: str,
    role: str = "end_user",
    *,
    password_hash: str = "",
    must_change_password: bool = True,
) -> None:
    """Insert a new user row. Raises if the email already exists."""
    if role not in ("admin", "end_user"):
        raise ValueError(f"invalid role: {role!r}")
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO users (email, password_hash, must_change_password, role) "
            "VALUES (?, ?, ?, ?)",
            (email.lower(), password_hash, 1 if must_change_password else 0, role),
        )


def set_user_password_hash(
    email: str,
    password_hash: str,
    *,
    must_change_password: bool,
) -> None:
    """Replace the stored password hash for a user."""
    with get_connection() as conn:
        conn.execute(
            "UPDATE users SET password_hash = ?, must_change_password = ?, "
            "updated_at = datetime('now') WHERE email = ?",
            (password_hash, 1 if must_change_password else 0, email.lower()),
        )


def set_user_role(email: str, role: str) -> None:
    """Update a user's role."""
    if role not in ("admin", "end_user"):
        raise ValueError(f"invalid role: {role!r}")
    with get_connection() as conn:
        conn.execute(
            "UPDATE users SET role = ?, updated_at = datetime('now') "
            "WHERE email = ?",
            (role, email.lower()),
        )


def delete_user(email: str) -> None:
    """Remove a user. No-op if the email isn't in the table."""
    with get_connection() as conn:
        conn.execute("DELETE FROM users WHERE email = ?", (email.lower(),))


# ---------------------------------------------------------------------------
# Rubrics
# ---------------------------------------------------------------------------

def save_rubric(rubric) -> int:
    """Insert a rubric and its sections. Return rubric_id.

    Args:
        rubric: A SyntheticRubric dataclass instance.
    """
    score_table = json.dumps(
        [(r, m) for r, m in rubric.score_table] if rubric.score_table else []
    )

    with get_connection() as conn:
        cur = conn.execute(
            "INSERT INTO rubrics "
            "(assessment_type_id, case_title, case_description, "
            " learner_instructions, model_answer, score_table_json, source) "
            "VALUES (?, ?, ?, ?, ?, ?, 'synthetic')",
            (
                rubric.assessment_type_id,
                rubric.case_title or "",
                rubric.case_description or "",
                rubric.learner_instructions or "",
                rubric.model_answer or "",
                score_table,
            ),
        )
        rubric_id = cur.lastrowid

        for sort_idx, (sec_key, sec) in enumerate(rubric.sections.items()):
            conn.execute(
                "INSERT INTO rubric_sections "
                "(rubric_id, section_key, display_name, max_score, criteria, "
                " score_levels_json, checklist_items_json, sort_order) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    rubric_id,
                    sec.section_key,
                    sec.display_name,
                    sec.max_score,
                    sec.criteria or "",
                    json.dumps(
                        {str(k): v for k, v in sec.score_levels.items()}
                        if sec.score_levels else {}
                    ),
                    json.dumps(sec.checklist_items if sec.checklist_items else []),
                    sort_idx,
                ),
            )

    logger.info("Saved rubric %d with %d sections", rubric_id, len(rubric.sections))
    return rubric_id


def get_rubric(rubric_id: int) -> Optional[dict]:
    """Return rubric as dict with sections, or None."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM rubrics WHERE rubric_id = ?", (rubric_id,)
        ).fetchone()
        if not row:
            return None

        sections = conn.execute(
            "SELECT * FROM rubric_sections WHERE rubric_id = ? ORDER BY sort_order",
            (rubric_id,),
        ).fetchall()

    result = dict(row)
    result["sections"] = [dict(s) for s in sections]
    return result


def get_rubric_sections_as_parsed(rubric_id: int) -> dict:
    """Return rubric sections in the format expected by build_user_message().

    Returns:
        {section_key: {"criteria": str, "max_score": float,
                       "checklist_items": list, "score_levels": dict}}
    """
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT section_key, display_name, max_score, criteria, "
            "       score_levels_json, checklist_items_json "
            "FROM rubric_sections WHERE rubric_id = ? ORDER BY sort_order",
            (rubric_id,),
        ).fetchall()

    parsed = {}
    for row in rows:
        checklist = json.loads(row["checklist_items_json"])
        score_levels = json.loads(row["score_levels_json"])

        # Build criteria text from structured data
        if checklist:
            lines = []
            for item in checklist:
                line = f"- {item['item']}: {item['points']} pt"
                if item.get("partial"):
                    line += f" (partial: 0.5 pt if {item['partial']})"
                lines.append(line)
            criteria_text = "\n".join(lines)
        elif score_levels:
            lines = []
            for lvl, desc in sorted(
                score_levels.items(), key=lambda x: int(x[0]), reverse=True
            ):
                lines.append(f"{lvl}: {desc}")
            criteria_text = "\n".join(lines)
        else:
            criteria_text = row["criteria"] or "No specific criteria available."

        parsed[row["section_key"]] = {
            "criteria": criteria_text,
            "max_score": row["max_score"],
            "checklist_items": checklist,
            "score_levels": score_levels,
        }

    return parsed


# ---------------------------------------------------------------------------
# Synthetic Sessions
# ---------------------------------------------------------------------------

def save_synthetic_session(session, rubric_id: int,
                           model_used: str = None,
                           params: dict = None) -> int:
    """Insert a synthetic session linked to a rubric. Return session_id.

    Args:
        session: A SyntheticSession dataclass instance.
        rubric_id: The rubric this session uses.
        model_used: LLM model identifier.
        params: Generation parameters dict.
    """
    from dataclasses import asdict

    faculty_dict = asdict(session.faculty) if session.faculty else {}
    students_list = [asdict(s) for s in session.students] if session.students else []

    # Convert student_notes keys to strings for JSON
    notes = {str(k): v for k, v in session.student_notes.items()}
    scores = {str(k): v for k, v in session.faculty_scores.items()}

    with get_connection() as conn:
        # Deactivate other sessions for this type
        conn.execute(
            "UPDATE synthetic_sessions SET is_active = 0 "
            "WHERE assessment_type_id = ?",
            (session.assessment_type_id,),
        )

        cur = conn.execute(
            "INSERT INTO synthetic_sessions "
            "(assessment_type_id, rubric_id, label, faculty_json, students_json, "
            " student_notes_json, faculty_scores_json, sections_json, "
            " model_used, generation_params_json, is_active) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)",
            (
                session.assessment_type_id,
                rubric_id,
                session.label,
                json.dumps(faculty_dict),
                json.dumps(students_list),
                json.dumps(notes),
                json.dumps(scores),
                json.dumps(session.sections),
                model_used or "",
                json.dumps(params or {}),
            ),
        )
        session_id = cur.lastrowid

    logger.info("Saved session %d for type %s", session_id, session.assessment_type_id)
    return session_id


def get_active_session(assessment_type_id: str) -> Optional[dict]:
    """Return the active session for a type, or None."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM synthetic_sessions "
            "WHERE assessment_type_id = ? AND is_active = 1 "
            "ORDER BY created_at DESC LIMIT 1",
            (assessment_type_id,),
        ).fetchone()
    return dict(row) if row else None


def set_active_session(assessment_type_id: str, session_id: int) -> None:
    """Mark one session as active, deactivate others for the type."""
    with get_connection() as conn:
        conn.execute(
            "UPDATE synthetic_sessions SET is_active = 0 "
            "WHERE assessment_type_id = ?",
            (assessment_type_id,),
        )
        conn.execute(
            "UPDATE synthetic_sessions SET is_active = 1 "
            "WHERE session_id = ?",
            (session_id,),
        )


def get_sessions_for_type(assessment_type_id: str) -> list[dict]:
    """Return all sessions for a type, ordered by created_at desc."""
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT session_id, label, is_active, created_at, model_used "
            "FROM synthetic_sessions "
            "WHERE assessment_type_id = ? "
            "ORDER BY created_at DESC",
            (assessment_type_id,),
        ).fetchall()
    return [dict(r) for r in rows]


def has_synthetic_data(assessment_type_id: str) -> bool:
    """Check if any session exists for this type."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT COUNT(*) FROM synthetic_sessions "
            "WHERE assessment_type_id = ?",
            (assessment_type_id,),
        ).fetchone()
    return row[0] > 0


def delete_synthetic_sessions(assessment_type_id: str) -> None:
    """Delete all sessions and their rubrics for a type."""
    with get_connection() as conn:
        # Get rubric IDs to clean up
        rubric_ids = conn.execute(
            "SELECT DISTINCT rubric_id FROM synthetic_sessions "
            "WHERE assessment_type_id = ?",
            (assessment_type_id,),
        ).fetchall()

        conn.execute(
            "DELETE FROM synthetic_sessions WHERE assessment_type_id = ?",
            (assessment_type_id,),
        )

        # Delete orphaned rubrics (only synthetic ones not used by other sessions)
        for row in rubric_ids:
            rid = row["rubric_id"]
            count = conn.execute(
                "SELECT COUNT(*) FROM synthetic_sessions WHERE rubric_id = ?",
                (rid,),
            ).fetchone()[0]
            if count == 0:
                conn.execute("DELETE FROM rubrics WHERE rubric_id = ?", (rid,))


def delete_session(session_id: int) -> None:
    """Delete one session (and orphaned rubric if applicable)."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT rubric_id FROM synthetic_sessions WHERE session_id = ?",
            (session_id,),
        ).fetchone()
        if not row:
            return

        rubric_id = row["rubric_id"]
        conn.execute(
            "DELETE FROM synthetic_sessions WHERE session_id = ?", (session_id,)
        )

        # Delete orphaned rubric
        count = conn.execute(
            "SELECT COUNT(*) FROM synthetic_sessions WHERE rubric_id = ?",
            (rubric_id,),
        ).fetchone()[0]
        if count == 0:
            conn.execute("DELETE FROM rubrics WHERE rubric_id = ?", (rubric_id,))


def get_session_data(session_id: int) -> Optional[dict]:
    """Return full session data including all JSON fields."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM synthetic_sessions WHERE session_id = ?",
            (session_id,),
        ).fetchone()
    if not row:
        return None
    result = dict(row)
    # Parse JSON fields
    result["faculty"] = json.loads(result["faculty_json"])
    result["students"] = json.loads(result["students_json"])
    result["student_notes"] = json.loads(result["student_notes_json"])
    result["faculty_scores"] = json.loads(result["faculty_scores_json"])
    result["sections"] = json.loads(result["sections_json"])
    return result


# ---------------------------------------------------------------------------
# Example Files
# ---------------------------------------------------------------------------

def save_example_file(assessment_type_id: str, slot: str,
                      filename: str, data: bytes) -> None:
    """Store or replace an example file as a BLOB."""
    with get_connection() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO example_files "
            "(assessment_type_id, slot, original_filename, file_data, uploaded_at) "
            "VALUES (?, ?, ?, ?, datetime('now'))",
            (assessment_type_id, slot, filename, data),
        )


def get_example_file(assessment_type_id: str,
                     slot: str) -> Optional[tuple[str, bytes]]:
    """Return (original_filename, file_data) or None."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT original_filename, file_data FROM example_files "
            "WHERE assessment_type_id = ? AND slot = ?",
            (assessment_type_id, slot),
        ).fetchone()
    return (row["original_filename"], row["file_data"]) if row else None


def get_example_display_name(assessment_type_id: str,
                             slot: str) -> Optional[str]:
    """Return just the original filename, or None."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT original_filename FROM example_files "
            "WHERE assessment_type_id = ? AND slot = ?",
            (assessment_type_id, slot),
        ).fetchone()
    return row["original_filename"] if row else None


def remove_example_file(assessment_type_id: str, slot: str) -> None:
    """Delete the example file for a slot."""
    with get_connection() as conn:
        conn.execute(
            "DELETE FROM example_files "
            "WHERE assessment_type_id = ? AND slot = ?",
            (assessment_type_id, slot),
        )


def write_example_to_temp(assessment_type_id: str, slot: str) -> Optional[str]:
    """Write a BLOB example file to a temp file and return the path.

    Returns None if no file exists for the given slot.
    Needed because rubric parsing and Excel loading expect file paths.
    """
    result = get_example_file(assessment_type_id, slot)
    if result is None:
        return None

    filename, data = result
    ext = os.path.splitext(filename)[1]
    fd, path = tempfile.mkstemp(suffix=ext)
    with os.fdopen(fd, "wb") as f:
        f.write(data)
    return path


# ---------------------------------------------------------------------------
# Grading Runs
# ---------------------------------------------------------------------------

def create_grading_run(
    assessment_type_id: str,
    model_used: str,
    temperature: float,
    source_type: str = "uploaded",
    rubric_id: Optional[int] = None,
    session_id: Optional[int] = None,
) -> int:
    """Insert a new grading run. Return run_id."""
    with get_connection() as conn:
        cur = conn.execute(
            "INSERT INTO grading_runs "
            "(assessment_type_id, rubric_id, model_used, temperature, "
            " source_type, session_id) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (assessment_type_id, rubric_id, model_used, temperature,
             source_type, session_id),
        )
    return cur.lastrowid


def update_grading_run(
    run_id: int,
    status: str,
    results_json: str = None,
    log_text: str = None,
) -> None:
    """Update a grading run's status and optionally store results."""
    with get_connection() as conn:
        conn.execute(
            "UPDATE grading_runs SET status = ?, results_json = ?, "
            "log_text = ?, completed_at = datetime('now') "
            "WHERE run_id = ?",
            (status, results_json, log_text, run_id),
        )


def get_latest_grading_run(assessment_type_id: str) -> Optional[dict]:
    """Return the most recent completed grading run, or None."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM grading_runs "
            "WHERE assessment_type_id = ? AND status = 'completed' "
            "ORDER BY completed_at DESC LIMIT 1",
            (assessment_type_id,),
        ).fetchone()
    return dict(row) if row else None


# ---------------------------------------------------------------------------
# Migration from flat files
# ---------------------------------------------------------------------------

def migrate_flat_files(repo_root: str) -> None:
    """One-time migration: import .env, synthetic_data/, and examples/ into DB.

    Idempotent — skips if data already exists in DB or source dirs are gone.
    Renames source directories to .bak after successful migration.
    """
    _migrate_env_file(repo_root)
    _migrate_synthetic_data(repo_root)
    _migrate_example_files(repo_root)


def _migrate_env_file(repo_root: str) -> None:
    """Import API keys from .env file."""
    env_path = os.path.join(repo_root, ".env")
    if not os.path.isfile(env_path):
        return

    # Check if we already have keys in DB
    existing = load_api_keys()
    if existing:
        return

    provider_map = {
        "OPENAI_API_KEY": "openai",
        "ANTHROPIC_API_KEY": "anthropic",
        "GOOGLE_API_KEY": "google",
    }

    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key in provider_map and value:
                save_api_key(provider_map[key], key, value)
                logger.info("Migrated API key: %s", key)


def _migrate_synthetic_data(repo_root: str) -> None:
    """Import synthetic data directories into DB."""
    synth_dir = os.path.join(repo_root, "synthetic_data")
    if not os.path.isdir(synth_dir):
        return

    # Check if we already have sessions in DB
    with get_connection() as conn:
        count = conn.execute("SELECT COUNT(*) FROM synthetic_sessions").fetchone()[0]
    if count > 0:
        return

    import pandas as pd

    for type_id in os.listdir(synth_dir):
        type_dir = os.path.join(synth_dir, type_id)
        if not os.path.isdir(type_dir):
            continue

        notes_path = os.path.join(type_dir, "student_notes.xlsx")
        if not os.path.isfile(notes_path):
            continue

        logger.info("Migrating synthetic data for type: %s", type_id)

        # Create a minimal rubric from rubric.txt if available
        rubric_text = ""
        rubric_txt_path = os.path.join(type_dir, "rubric.txt")
        if os.path.isfile(rubric_txt_path):
            with open(rubric_txt_path) as f:
                rubric_text = f.read()

        with get_connection() as conn:
            cur = conn.execute(
                "INSERT INTO rubrics (assessment_type_id, case_title, "
                " learner_instructions, source) "
                "VALUES (?, 'Migrated Rubric', ?, 'synthetic')",
                (type_id, rubric_text),
            )
            rubric_id = cur.lastrowid

            # Try to determine sections from the type
            sections_map = {
                "kpsom_ipass": {
                    "illness_severity": ("Illness Severity", 2),
                    "patient_summary": ("Patient Summary", 14),
                    "action_list": ("Action List", 5),
                    "situation_awareness": ("Situation Awareness", 3),
                    "organization": ("Organization", 3),
                },
                "kpsom_documentation": {
                    "hpi": ("HPI", 5),
                    "social_history": ("Social History", 5),
                    "summary_statement": ("Summary Statement", 5),
                    "assessment": ("Assessment", 5),
                    "plan": ("Plan", 5),
                    "written_communication": ("Written Communication", 4),
                },
            }
            type_sections = sections_map.get(type_id, {})
            for idx, (sec_key, (display, max_sc)) in enumerate(type_sections.items()):
                conn.execute(
                    "INSERT INTO rubric_sections "
                    "(rubric_id, section_key, display_name, max_score, "
                    " criteria, sort_order) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (rubric_id, sec_key, display, max_sc,
                     "Migrated from flat files — regenerate for full criteria.",
                     idx),
                )

        # Read student notes
        try:
            raw = pd.read_excel(notes_path, header=None)
            if raw.shape[0] >= 2:
                headers = raw.iloc[1].tolist()
                headers = [
                    str(h).strip() if pd.notna(h) else f"col_{i}"
                    for i, h in enumerate(headers)
                ]
                data = raw.iloc[2:].reset_index(drop=True)
                data.columns = headers

                student_notes = {}
                for _, row in data.iterrows():
                    sid = str(row.get("Student", row.get("Student ID", "")))
                    if not sid:
                        continue
                    notes = {}
                    for col in headers:
                        if col.lower() not in ("student", "student id") and pd.notna(
                            row.get(col)
                        ):
                            notes[col] = str(row[col])
                    student_notes[sid] = notes
            else:
                student_notes = {}
        except Exception as exc:
            logger.warning("Could not read student notes for %s: %s", type_id, exc)
            student_notes = {}

        # Read faculty scores
        faculty_scores = {}
        scores_path = os.path.join(type_dir, "faculty_scores.xlsx")
        if os.path.isfile(scores_path):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(scores_path, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if len(all_rows) >= 3:
                    fheaders = [
                        str(h).strip() if h else f"col_{i}"
                        for i, h in enumerate(all_rows[2])
                    ]
                    for drow in all_rows[3:]:
                        row_dict = dict(zip(fheaders, drow))
                        sid = row_dict.get("Student", row_dict.get("Student ID"))
                        if sid is None:
                            continue
                        scores = {}
                        for k, v in row_dict.items():
                            if k.lower() not in ("student", "student id") and v is not None:
                                try:
                                    scores[k] = float(v)
                                except (ValueError, TypeError):
                                    pass
                        faculty_scores[str(int(float(sid)))] = scores
            except Exception as exc:
                logger.warning("Could not read faculty scores for %s: %s", type_id, exc)

        sections = list(type_sections.keys()) if type_sections else []

        with get_connection() as conn:
            conn.execute(
                "INSERT INTO synthetic_sessions "
                "(assessment_type_id, rubric_id, label, faculty_json, "
                " students_json, student_notes_json, faculty_scores_json, "
                " sections_json, model_used, is_active) "
                "VALUES (?, ?, 'Migrated Session', '{}', '[]', ?, ?, ?, '', 1)",
                (
                    type_id,
                    rubric_id,
                    json.dumps(student_notes),
                    json.dumps(faculty_scores),
                    json.dumps(sections),
                ),
            )

        logger.info("Migrated synthetic session for %s", type_id)

    # Rename to .bak
    bak = synth_dir + ".bak"
    if not os.path.exists(bak):
        os.rename(synth_dir, bak)
        logger.info("Renamed %s to %s", synth_dir, bak)


def _migrate_example_files(repo_root: str) -> None:
    """Import example file directories into DB."""
    examples_dir = os.path.join(repo_root, "examples")
    if not os.path.isdir(examples_dir):
        return

    # Check if we already have examples in DB
    with get_connection() as conn:
        count = conn.execute("SELECT COUNT(*) FROM example_files").fetchone()[0]
    if count > 0:
        return

    for type_id in os.listdir(examples_dir):
        type_dir = os.path.join(examples_dir, type_id)
        if not os.path.isdir(type_dir):
            continue

        for fname in os.listdir(type_dir):
            if "__" not in fname:
                continue
            slot, orig_name = fname.split("__", 1)
            file_path = os.path.join(type_dir, fname)
            with open(file_path, "rb") as f:
                save_example_file(type_id, slot, orig_name, f.read())
            logger.info("Migrated example file: %s/%s", type_id, fname)

    # Rename to .bak
    bak = examples_dir + ".bak"
    if not os.path.exists(bak):
        os.rename(examples_dir, bak)
        logger.info("Renamed %s to %s", examples_dir, bak)

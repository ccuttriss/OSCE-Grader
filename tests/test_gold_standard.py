"""Tests for gold standard faculty bias analysis module."""

import json
import os
import sys

import pandas as pd
import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

from gold_standard import (
    BiasAnalysisResult,
    ConsensusResult,
    CrossSessionStats,
    GoldStandardBenchmark,
    SectionStats,
    SessionData,
    build_bias_prompt,
    compute_consensus_analysis,
    compute_cross_session_stats,
    generate_benchmark_excel,
    generate_benchmark_json,
    load_faculty_session,
    parse_bias_response,
    validate_sessions,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_session(label, type_id, scores):
    """Helper to create a SessionData from a scores dict."""
    from gold_standard import _SECTION_KEYS

    sections = _SECTION_KEYS[type_id]
    return SessionData(
        label=label,
        assessment_type_id=type_id,
        sections=list(sections),
        scores=scores,
        student_count=len(scores),
    )


@pytest.fixture
def ipass_faculty_xlsx(tmp_path):
    """Create a minimal I-PASS faculty scores Excel (3 header rows + data)."""
    wb = Workbook()
    ws = wb.active
    # Row 0: group header (skipped)
    ws.append(["", "Group A", "", "", "", ""])
    # Row 1: sub-group header (skipped)
    ws.append(["", "Sub A", "", "", "", ""])
    # Row 2: real headers
    ws.append([
        "Student",
        "Illness Severity",
        "Patient Summary",
        "Action List",
        "Situation Awareness",
        "Organization",
    ])
    # Row 3+: data
    ws.append([1, 1.5, 10.0, 3.0, 2.0, 2.5])
    ws.append([2, 2.0, 12.0, 4.0, 3.0, 3.0])
    ws.append([3, 1.0, 8.0, 2.5, 1.5, 1.0])

    path = str(tmp_path / "ipass_faculty.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def documentation_faculty_xlsx(tmp_path):
    """Create a minimal documentation faculty scores Excel.

    Row 0 = flat headers. Scores at column indices 29,34,49,58,72,73.
    """
    wb = Workbook()
    ws = wb.active

    # Build header row with enough columns (need at least 74)
    headers = [""] * 82
    headers[29] = "HPI"
    headers[34] = "Social Hx"
    headers[49] = "Summary Statement"
    headers[58] = "Assessment"
    headers[72] = "Plan"
    headers[73] = "Org/Lang"
    headers[74] = "PCIG Total"
    headers[75] = "PCIG Milestone"
    headers[76] = "PCDP Total"
    headers[77] = "PCDP Milestone"
    headers[78] = "PCDO Score"
    headers[79] = "PCDO Milestone"
    headers[80] = "Total"
    headers[81] = "Total Milestone"
    ws.append(headers)

    # Data rows
    for student_scores in [(3, 4, 3, 4, 3, 2), (5, 5, 4, 5, 4, 3)]:
        row = [None] * 82
        row[29] = student_scores[0]
        row[34] = student_scores[1]
        row[49] = student_scores[2]
        row[58] = student_scores[3]
        row[72] = student_scores[4]
        row[73] = student_scores[5]
        ws.append(row)

    path = str(tmp_path / "doc_faculty.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def ethics_faculty_xlsx(tmp_path):
    """Create a minimal ethics faculty scores Excel.

    Row 0 = section groups, Row 1 = sub-items, Row 2+ = data.
    Key cols: 0=StudentID, 2=q1_problematic, 3=q1_reasonable, 4=q1_total,
    6=q2a, 8=q2b, 10=q2c, 11=q2_total, 16=q3_total, 17=task_total, 18=milestone.
    """
    wb = Workbook()
    ws = wb.active

    # Two header rows
    ws.append(["", "Q1", "", "", "", "Q2", "", "", "", "", "", "", "", "", "", "", "Q3", "", ""])
    ws.append(["ID", "", "prob", "reas", "total", "", "q2a", "", "q2b", "", "q2c", "q2tot", "", "", "", "", "q3tot", "task", "mile"])

    # Data rows
    ws.append([1, None, 1.5, 1.0, 2.5, None, 1.5, None, 2.0, None, 1.0, 4.5, None, None, None, None, 6.0, 13.0, "Mid-Developing"])
    ws.append([2, None, 2.0, 2.0, 4.0, None, 2.0, None, 2.0, None, 2.0, 6.0, None, None, None, None, 8.0, 18.0, "Aspirational"])

    path = str(tmp_path / "ethics_faculty.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Session loading tests
# ---------------------------------------------------------------------------


class TestLoadFacultySession:
    def test_load_ipass(self, ipass_faculty_xlsx):
        session = load_faculty_session(ipass_faculty_xlsx, "kpsom_ipass", "2023")
        assert session.label == "2023"
        assert session.assessment_type_id == "kpsom_ipass"
        assert session.student_count == 3
        assert 1 in session.scores
        assert session.scores[1]["illness_severity"] == 1.5
        assert session.scores[2]["patient_summary"] == 12.0

    def test_load_documentation(self, documentation_faculty_xlsx):
        session = load_faculty_session(
            documentation_faculty_xlsx, "kpsom_documentation", "2024"
        )
        assert session.student_count == 2
        # Student 1 (row index 0, 1-indexed)
        assert session.scores[1]["hpi"] == 3.0
        assert session.scores[1]["social_hx"] == 4.0
        assert session.scores[1]["org_lang"] == 2.0

    def test_load_ethics(self, ethics_faculty_xlsx):
        session = load_faculty_session(
            ethics_faculty_xlsx, "kpsom_ethics", "2025"
        )
        assert session.student_count == 2
        assert session.scores[1]["q1_total"] == 2.5
        assert session.scores[2]["q3_total"] == 8.0

    def test_unknown_type_raises(self, tmp_path):
        with pytest.raises(ValueError, match="Unknown assessment type"):
            load_faculty_session(str(tmp_path / "x.xlsx"), "bogus_type", "x")


# ---------------------------------------------------------------------------
# Validation tests
# ---------------------------------------------------------------------------


class TestValidation:
    def test_valid_two_sessions(self):
        s1 = _make_session("a", "kpsom_ipass", {1: {}})
        s2 = _make_session("b", "kpsom_ipass", {2: {}})
        assert validate_sessions([s1, s2]) is None

    def test_single_session_rejected(self):
        s = _make_session("a", "kpsom_ipass", {1: {}})
        err = validate_sessions([s])
        assert "At least 2" in err

    def test_too_many_sessions_rejected(self):
        sessions = [_make_session(f"s{i}", "kpsom_ipass", {}) for i in range(11)]
        err = validate_sessions(sessions)
        assert "Maximum of 10" in err

    def test_mismatched_types_rejected(self):
        s1 = _make_session("a", "kpsom_ipass", {})
        s2 = _make_session("b", "kpsom_ethics", {})
        err = validate_sessions([s1, s2])
        assert "same assessment type" in err


# ---------------------------------------------------------------------------
# Statistics tests
# ---------------------------------------------------------------------------


class TestComputeStats:
    def test_basic_stats(self):
        s1 = _make_session(
            "2023",
            "kpsom_documentation",
            {
                1: {"hpi": 3, "social_hx": 4, "summary_statement": 3, "assessment": 4, "plan": 3, "org_lang": 2},
                2: {"hpi": 5, "social_hx": 5, "summary_statement": 4, "assessment": 5, "plan": 4, "org_lang": 3},
            },
        )
        s2 = _make_session(
            "2024",
            "kpsom_documentation",
            {
                1: {"hpi": 4, "social_hx": 4, "summary_statement": 4, "assessment": 4, "plan": 4, "org_lang": 3},
                2: {"hpi": 2, "social_hx": 3, "summary_statement": 2, "assessment": 3, "plan": 2, "org_lang": 1},
            },
        )

        stats = compute_cross_session_stats([s1, s2])
        assert stats.session_count == 2
        assert stats.total_students == 4

        hpi = stats.section_stats["hpi"]
        # Values: 3, 5, 4, 2 → mean=3.5
        assert hpi.mean == 3.5
        assert hpi.min_score == 2
        assert hpi.max_score == 5
        assert "2023" in hpi.per_session_means
        assert hpi.per_session_means["2023"] == 4.0  # (3+5)/2
        assert hpi.per_session_means["2024"] == 3.0  # (4+2)/2

    def test_handles_none_values(self):
        s1 = _make_session(
            "2023",
            "kpsom_ethics",
            {
                1: {"q1_total": 3.0, "q2a_score": None, "q2b_score": 1.0, "q2c_score": 1.5, "q3_total": 6.0},
            },
        )
        s2 = _make_session(
            "2024",
            "kpsom_ethics",
            {
                1: {"q1_total": 4.0, "q2a_score": 2.0, "q2b_score": 1.5, "q2c_score": 2.0, "q3_total": 7.0},
            },
        )

        stats = compute_cross_session_stats([s1, s2])
        # q2a has only one value (2.0) since the other is None
        assert stats.section_stats["q2a_score"].mean == 2.0

    def test_empty_sessions_raises(self):
        with pytest.raises(ValueError, match="No sessions"):
            compute_cross_session_stats([])


# ---------------------------------------------------------------------------
# Bias prompt tests
# ---------------------------------------------------------------------------


class TestBuildBiasPrompt:
    def test_contains_type_name(self):
        stats = CrossSessionStats(
            assessment_type_id="kpsom_ipass",
            sections=["illness_severity"],
            section_stats={
                "illness_severity": SectionStats(
                    section="illness_severity",
                    mean=1.5,
                    median=1.5,
                    std=0.5,
                    min_score=1.0,
                    max_score=2.0,
                )
            },
            total_students=10,
            session_count=2,
            session_labels=["2023", "2024"],
        )
        messages = build_bias_prompt(stats, "kpsom_ipass")
        assert len(messages) == 2
        assert "I-PASS Handoff" in messages[0]["content"]
        assert "illness_severity" in messages[1]["content"]
        assert "mean=1.50" in messages[1]["content"]

    def test_all_sections_present(self):
        sections = ["hpi", "social_hx", "summary_statement", "assessment", "plan", "org_lang"]
        stats = CrossSessionStats(
            assessment_type_id="kpsom_documentation",
            sections=sections,
            section_stats={
                sec: SectionStats(
                    section=sec, mean=3.0, median=3.0, std=1.0,
                    min_score=1.0, max_score=5.0,
                )
                for sec in sections
            },
            total_students=20,
            session_count=3,
            session_labels=["2022", "2023", "2024"],
        )
        messages = build_bias_prompt(stats, "kpsom_documentation")
        user_msg = messages[1]["content"]
        for sec in sections:
            assert sec in user_msg


# ---------------------------------------------------------------------------
# Response parsing tests
# ---------------------------------------------------------------------------


class TestParseBiasResponse:
    def test_valid_json(self):
        raw = json.dumps({
            "systematic_biases": [
                {"section": "hpi", "direction": "lenient", "magnitude": "medium",
                 "description": "HPI scores cluster near ceiling."}
            ],
            "distribution_anomalies": [],
            "section_patterns": [],
            "drift_over_years": [],
            "outlier_patterns": [],
            "summary": "Overall scoring shows leniency.",
            "recommendations": ["Tighten HPI rubric."],
        })
        result = parse_bias_response(raw)
        assert len(result.systematic_biases) == 1
        assert result.systematic_biases[0]["section"] == "hpi"
        assert result.summary == "Overall scoring shows leniency."
        assert len(result.recommendations) == 1

    def test_json_with_fences(self):
        raw = "```json\n" + json.dumps({
            "systematic_biases": [],
            "distribution_anomalies": [],
            "section_patterns": [],
            "drift_over_years": [],
            "outlier_patterns": [],
            "summary": "OK",
            "recommendations": [],
        }) + "\n```"
        result = parse_bias_response(raw)
        assert result.summary == "OK"

    def test_malformed_json_fallback(self):
        result = parse_bias_response("This is not JSON at all.")
        assert "Failed to parse" in result.summary
        assert result.systematic_biases == []

    def test_missing_fields_default_to_empty(self):
        raw = json.dumps({"summary": "partial"})
        result = parse_bias_response(raw)
        assert result.summary == "partial"
        assert result.systematic_biases == []
        assert result.recommendations == []


# ---------------------------------------------------------------------------
# Benchmark generation tests
# ---------------------------------------------------------------------------


class TestGenerateBenchmark:
    def _make_benchmark(self):
        return GoldStandardBenchmark(
            assessment_type_id="kpsom_documentation",
            created_date="2026-03-06",
            sections={
                "hpi": {"min": 3.0, "max": 5.0, "notes": "Strong performance expected", "approved_by": "Dr. Smith"},
                "plan": {"min": 2.0, "max": 4.0, "notes": "", "approved_by": "Dr. Jones"},
            },
            bias_findings=BiasAnalysisResult(
                summary="Minor leniency detected.",
                recommendations=["Calibrate Plan scoring."],
            ),
            stats=CrossSessionStats(
                assessment_type_id="kpsom_documentation",
                sections=["hpi", "plan"],
                section_stats={
                    "hpi": SectionStats("hpi", 4.0, 4.0, 0.8, 2.0, 5.0,
                                        per_session_means={"2023": 3.8, "2024": 4.2}),
                    "plan": SectionStats("plan", 3.0, 3.0, 1.0, 1.0, 5.0,
                                         per_session_means={"2023": 2.8, "2024": 3.2}),
                },
                total_students=60,
                session_count=2,
                session_labels=["2023", "2024"],
            ),
            version=1,
        )

    def test_excel_has_three_sheets(self):
        bm = self._make_benchmark()
        data = generate_benchmark_excel(bm)
        assert len(data) > 0

        # Read back
        import io
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(io.BytesIO(data))
        assert wb.sheetnames == ["Benchmarks", "Bias Findings", "Statistics"]

    def test_excel_benchmarks_sheet(self):
        bm = self._make_benchmark()
        data = generate_benchmark_excel(bm)

        import io
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(io.BytesIO(data))
        ws = wb["Benchmarks"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("Section", "Benchmark Min", "Benchmark Max", "Notes", "Approved By")
        assert rows[1][0] == "hpi"
        assert rows[1][1] == 3.0
        assert rows[2][0] == "plan"

    def test_json_has_version(self):
        bm = self._make_benchmark()
        raw = generate_benchmark_json(bm)
        data = json.loads(raw)
        assert data["version"] == 1
        assert data["assessment_type_id"] == "kpsom_documentation"
        assert "hpi" in data["sections"]
        assert "statistics" in data
        assert data["statistics"]["sections"]["hpi"]["mean"] == 4.0

    def test_json_roundtrip(self):
        bm = self._make_benchmark()
        raw = generate_benchmark_json(bm)
        data = json.loads(raw)
        # Verify structure
        assert set(data["sections"].keys()) == {"hpi", "plan"}
        assert data["bias_findings"]["summary"] == "Minor leniency detected."
        assert len(data["bias_findings"]["recommendations"]) == 1


# ---------------------------------------------------------------------------
# Consensus analysis tests
# ---------------------------------------------------------------------------


def _make_sessions_with_means(type_id, means_per_session):
    """Create SessionData objects whose scores produce desired per-section means.

    means_per_session: dict[label → dict[section → target_mean]]
    Creates 2 students per session centered on the target mean.
    """
    from gold_standard import _SECTION_KEYS

    sections = _SECTION_KEYS[type_id]
    sessions = []
    for label, sec_means in means_per_session.items():
        scores = {}
        for i, offset in enumerate([-0.25, 0.25]):
            sid = i + 1
            student_scores = {}
            for sec in sections:
                target = sec_means.get(sec, 3.0)
                student_scores[sec] = target + offset
            scores[sid] = student_scores
        sessions.append(
            SessionData(
                label=label,
                assessment_type_id=type_id,
                sections=list(sections),
                scores=scores,
                student_count=2,
            )
        )
    return sessions


class TestConsensusAnalysis:
    def test_basic_three_sessions(self):
        sessions = _make_sessions_with_means("kpsom_documentation", {
            "2023": {"hpi": 3.0, "social_hx": 4.0, "summary_statement": 3.5, "assessment": 4.0, "plan": 3.0, "org_lang": 2.0},
            "2024": {"hpi": 3.2, "social_hx": 4.1, "summary_statement": 3.6, "assessment": 4.2, "plan": 3.1, "org_lang": 2.1},
            "2025": {"hpi": 2.0, "social_hx": 2.5, "summary_statement": 2.0, "assessment": 2.5, "plan": 2.0, "org_lang": 1.0},
        })
        stats = compute_cross_session_stats(sessions)
        result = compute_consensus_analysis(sessions, stats)

        # Basic structure checks
        assert len(result.eigenvalues) == 3
        assert result.eigenvalue_ratio > 0
        assert isinstance(result.single_culture_holds, bool)
        assert len(result.session_competence) == 3
        assert len(result.first_factor_loadings) == 3
        assert len(result.consensus_means) == 6  # 6 sections

        # Competence weights sum to ~1
        assert abs(sum(result.session_competence.values()) - 1.0) < 1e-6

        # Answer key has all sections
        for sec in ["hpi", "social_hx", "summary_statement", "assessment", "plan", "org_lang"]:
            assert sec in result.consensus_means
            assert sec in result.simple_means
            assert sec in result.divergence

    def test_two_sessions_edge_case(self):
        sessions = _make_sessions_with_means("kpsom_ethics", {
            "2023": {"q1_total": 3.0, "q2a_score": 1.5, "q2b_score": 1.0, "q2c_score": 1.5, "q3_total": 6.0},
            "2024": {"q1_total": 3.5, "q2a_score": 1.8, "q2b_score": 1.2, "q2c_score": 1.6, "q3_total": 6.5},
        })
        stats = compute_cross_session_stats(sessions)
        result = compute_consensus_analysis(sessions, stats)

        assert "2 sessions" in result.fit_label
        assert len(result.eigenvalues) == 2
        assert len(result.session_competence) == 2

    def test_perfect_agreement(self):
        """Three identical sessions should yield equal competence and consensus == simple mean."""
        means = {"hpi": 3.0, "social_hx": 4.0, "summary_statement": 3.5, "assessment": 4.0, "plan": 3.0, "org_lang": 2.0}
        sessions = _make_sessions_with_means("kpsom_documentation", {
            "2023": means,
            "2024": means,
            "2025": means,
        })
        stats = compute_cross_session_stats(sessions)
        result = compute_consensus_analysis(sessions, stats)

        # All competence scores should be approximately equal
        weights = list(result.session_competence.values())
        assert abs(max(weights) - min(weights)) < 0.01

        # Consensus should equal simple mean (no divergence)
        for sec in result.consensus_means:
            assert abs(result.consensus_means[sec] - result.simple_means[sec]) < 0.01

    def test_consensus_differs_from_simple_mean(self):
        """When 2 sessions agree and 1 diverges, consensus should favor the agreeing pair."""
        sessions = _make_sessions_with_means("kpsom_documentation", {
            "2023": {"hpi": 4.0, "social_hx": 4.0, "summary_statement": 4.0, "assessment": 4.0, "plan": 4.0, "org_lang": 3.0},
            "2024": {"hpi": 4.0, "social_hx": 4.0, "summary_statement": 4.0, "assessment": 4.0, "plan": 4.0, "org_lang": 3.0},
            "2025": {"hpi": 1.0, "social_hx": 1.0, "summary_statement": 1.0, "assessment": 1.0, "plan": 1.0, "org_lang": 1.0},
        })
        stats = compute_cross_session_stats(sessions)
        result = compute_consensus_analysis(sessions, stats)

        # 2025 should have lower competence than 2023/2024
        assert result.session_competence["2025"] < result.session_competence["2023"]

        # Consensus mean should be closer to 4.0 than the simple mean of 3.0
        assert result.consensus_means["hpi"] > result.simple_means["hpi"]

    def test_eigenvalues_are_exposed(self):
        """Eigenvalues should be available for inspection."""
        sessions = _make_sessions_with_means("kpsom_documentation", {
            "2023": {"hpi": 3.0, "social_hx": 4.0, "summary_statement": 3.5, "assessment": 4.0, "plan": 3.0, "org_lang": 2.0},
            "2024": {"hpi": 3.1, "social_hx": 4.1, "summary_statement": 3.6, "assessment": 4.1, "plan": 3.1, "org_lang": 2.1},
            "2025": {"hpi": 2.0, "social_hx": 2.5, "summary_statement": 2.0, "assessment": 2.5, "plan": 2.0, "org_lang": 1.0},
        })
        stats = compute_cross_session_stats(sessions)
        result = compute_consensus_analysis(sessions, stats)

        # Eigenvalues should be in descending order
        for i in range(len(result.eigenvalues) - 1):
            assert result.eigenvalues[i] >= result.eigenvalues[i + 1]

        # First eigenvalue should be the largest
        assert result.eigenvalues[0] == max(result.eigenvalues)

        # Ratio should be first / second (unless second is near zero)
        if abs(result.eigenvalues[1]) > 1e-10:
            expected_ratio = result.eigenvalues[0] / abs(result.eigenvalues[1])
            assert abs(result.eigenvalue_ratio - expected_ratio) < 0.01
        else:
            assert result.eigenvalue_ratio == float("inf")

    def test_first_factor_loadings_signed(self):
        """First-factor loadings should be signed (not just absolute values)."""
        sessions = _make_sessions_with_means("kpsom_documentation", {
            "2023": {"hpi": 4.0, "social_hx": 4.0, "summary_statement": 4.0, "assessment": 4.0, "plan": 4.0, "org_lang": 3.0},
            "2024": {"hpi": 4.0, "social_hx": 4.0, "summary_statement": 4.0, "assessment": 4.0, "plan": 4.0, "org_lang": 3.0},
            "2025": {"hpi": 1.0, "social_hx": 1.0, "summary_statement": 1.0, "assessment": 1.0, "plan": 1.0, "org_lang": 1.0},
        })
        stats = compute_cross_session_stats(sessions)
        result = compute_consensus_analysis(sessions, stats)

        # Loadings should be actual signed floats
        for label, loading in result.first_factor_loadings.items():
            assert isinstance(loading, float)

        # The agreeing sessions should have same-sign loadings
        assert (
            result.first_factor_loadings["2023"]
            * result.first_factor_loadings["2024"]
            > 0
        ), "Agreeing sessions should have same-sign loadings"

    def test_consensus_in_benchmark_json(self):
        """Consensus data should appear in the benchmark JSON."""
        sessions = _make_sessions_with_means("kpsom_documentation", {
            "2023": {"hpi": 3.0, "social_hx": 4.0, "summary_statement": 3.0, "assessment": 4.0, "plan": 3.0, "org_lang": 2.0},
            "2024": {"hpi": 3.5, "social_hx": 4.5, "summary_statement": 3.5, "assessment": 4.5, "plan": 3.5, "org_lang": 2.5},
        })
        stats = compute_cross_session_stats(sessions)
        consensus = compute_consensus_analysis(sessions, stats)

        bm = GoldStandardBenchmark(
            assessment_type_id="kpsom_documentation",
            created_date="2026-03-06",
            consensus=consensus,
        )
        raw = generate_benchmark_json(bm)
        data = json.loads(raw)

        assert "consensus" in data
        c = data["consensus"]
        assert "eigenvalues" in c
        assert "eigenvalue_ratio" in c
        assert "first_factor_loadings" in c
        assert "session_competence" in c
        assert "consensus_means" in c
        assert "simple_means" in c
        assert "divergence" in c
        assert "has_negative_loadings" in c

    def test_consensus_in_benchmark_excel(self):
        """Consensus data should appear as a 4th sheet in the Excel output."""
        sessions = _make_sessions_with_means("kpsom_documentation", {
            "2023": {"hpi": 3.0, "social_hx": 4.0, "summary_statement": 3.0, "assessment": 4.0, "plan": 3.0, "org_lang": 2.0},
            "2024": {"hpi": 3.5, "social_hx": 4.5, "summary_statement": 3.5, "assessment": 4.5, "plan": 3.5, "org_lang": 2.5},
        })
        stats = compute_cross_session_stats(sessions)
        consensus = compute_consensus_analysis(sessions, stats)

        bm = GoldStandardBenchmark(
            assessment_type_id="kpsom_documentation",
            created_date="2026-03-06",
            consensus=consensus,
        )
        data = generate_benchmark_excel(bm)

        import io
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(io.BytesIO(data))
        assert "Consensus Analysis" in wb.sheetnames

        ws = wb["Consensus Analysis"]
        rows = list(ws.iter_rows(values_only=True))
        # Check key content
        texts = [str(cell) for row in rows for cell in row if cell is not None]
        assert "Eigenvalue Ratio" in texts
        assert "Consensus Answer Key" in texts
        assert "First-Factor Loading (signed)" in texts
